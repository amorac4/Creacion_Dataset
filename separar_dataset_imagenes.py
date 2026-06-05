#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Separar Dataset de Imagenes
===========================

Genera datasets estructurados de imagenes a partir de Dataset_V1, la propuesta
temporal en Excel y las imagenes generadas en Lab_Creacion_Dataset.

Salida por combinacion:

    <dataset_name>-<balance_index>-<algorithm_index>__<balance>__<algorithm>/
      train/<familia>/<hash>.png
      val/<familia>/<hash>.png
      test/<familia>/<hash>.png

Tambien genera manifest y resumen para auditar hashes copiados/faltantes.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import shutil
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_CONFIG_PATH = Path("configs_dataset/config_separador_imagenes.json")

DEFAULT_CONFIG: dict[str, Any] = {
    "dataset_name": "Dataset_V1",
    "dataset_csv_dir": "Dataset_V1/csv",
    "split_workbook": "Dataset_V1/Analisis/propuesta_estructuracion_temporal_tablas.xlsx",
    "source_images_root": "C:/Users/ADOLF/Desktop/Repositorios/Lab_Creacion_Dataset/data",
    "source_collections": ["img"],
    "output_dir": "outputs/datasets_imagenes",
    "balances": ["Split_Todo", "Split_1000", "Split_1500", "Split_2000"],
    "algorithms": [
        {"name": "rgb", "source": "bin2rgb"},
        {"name": "simhash", "source": "simhash"},
        {"name": "dct", "source": "bigram_dct"},
        {"name": "markov", "source": "markov"},
        {"name": "wem", "source": "wem"},
    ],
    "valid_year_from": 1992,
    "valid_year_to": 2026,
    "split_folder_names": {
        "train": "train",
        "validacion": "val",
        "test": "test",
    },
    "folder_template": "{dataset_name}-{balance_index}-{algorithm_index}__{balance_slug}__{algorithm}",
    "image_extensions": [".png", ".jpg", ".jpeg", ".webp"],
    "copy_workers": 0,
    "overwrite": False,
}


@dataclass(frozen=True)
class DatasetRow:
    hash_md5: str
    family: str
    creation_day: dt.date
    source_csv: str
    original: dict[str, Any]


@dataclass(frozen=True)
class FamilySplitCounts:
    family: str
    total: int
    train: int
    validation: int
    test: int


@dataclass(frozen=True)
class BalanceSpec:
    sheet: str
    name: str
    families: dict[str, FamilySplitCounts]


@dataclass(frozen=True)
class AlgorithmSpec:
    name: str
    source: str


@dataclass(frozen=True)
class Assignment:
    balance: BalanceSpec
    row: DatasetRow
    split: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Copia imagenes por balance, algoritmo, split y familia."
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH, help="JSON de configuracion.")
    parser.add_argument("--dataset-name", help="Nombre base para las carpetas de salida.")
    parser.add_argument("--output-dir", type=Path, help="Directorio de salida.")
    parser.add_argument("--source-images-root", type=Path, help="Raiz de imagenes generadas.")
    parser.add_argument(
        "--balances",
        nargs="+",
        help="Hojas de balance a procesar, por ejemplo Split_1500 Split_2000.",
    )
    parser.add_argument(
        "--algorithms",
        nargs="+",
        help="Algoritmos por nombre de salida o source, por ejemplo rgb simhash dct markov.",
    )
    parser.add_argument("--dry-run", action="store_true", help="No copia imagenes; solo calcula y reporta.")
    parser.add_argument("--overwrite", action="store_true", help="Sobrescribe imagenes existentes.")
    return parser.parse_args()


def deep_update(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    result = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_update(result[key], value)
        else:
            result[key] = value
    return result


def load_config(path: Path) -> dict[str, Any]:
    if path.exists():
        user_config = json.loads(path.read_text(encoding="utf-8-sig"))
        return deep_update(DEFAULT_CONFIG, user_config)
    return dict(DEFAULT_CONFIG)


def apply_cli_overrides(config: dict[str, Any], args: argparse.Namespace) -> dict[str, Any]:
    if args.dataset_name:
        config["dataset_name"] = args.dataset_name
    if args.output_dir:
        config["output_dir"] = str(args.output_dir)
    if args.source_images_root:
        config["source_images_root"] = str(args.source_images_root)
    if args.balances:
        config["balances"] = args.balances
    if args.overwrite:
        config["overwrite"] = True
    if args.algorithms:
        selected = set(args.algorithms)
        configured = parse_algorithms(config)
        filtered = [
            {"name": algorithm.name, "source": algorithm.source}
            for algorithm in configured
            if algorithm.name in selected or algorithm.source in selected
        ]
        unknown = sorted(selected - {item["name"] for item in filtered} - {item["source"] for item in filtered})
        if unknown:
            raise SystemExit("Algoritmos no configurados: " + ", ".join(unknown))
        config["algorithms"] = filtered
    return config


def safe_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    output = []
    for char in text:
        if char.isalnum() or char in "._-":
            output.append(char)
        else:
            output.append("_")
    cleaned = "".join(output).strip("._-")
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned or "sin_valor"


def parse_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    if "T" in text:
        try:
            return dt.datetime.fromisoformat(text).date()
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def is_valid_temporal_date(date_value: dt.date | None, config: dict[str, Any]) -> bool:
    if date_value is None:
        return False
    year_from = int(config.get("valid_year_from") or 1992)
    year_to = int(config.get("valid_year_to") or 2026)
    return year_from <= date_value.year <= year_to


def load_dataset_rows(config: dict[str, Any]) -> list[DatasetRow]:
    csv_dir = Path(str(config["dataset_csv_dir"]))
    if not csv_dir.exists():
        raise SystemExit(f"No existe dataset_csv_dir: {csv_dir}")

    rows: list[DatasetRow] = []
    for csv_path in sorted(csv_dir.glob("*.csv")):
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for raw_row in reader:
                hash_md5 = str(raw_row.get("hash_md5") or "").strip().lower()
                family = safe_name(raw_row.get("familia_probable") or csv_path.stem)
                creation_day = parse_date(raw_row.get("dia_creacion_archivo") or raw_row.get("fecha_creacion_archivo"))
                if not hash_md5 or not is_valid_temporal_date(creation_day, config):
                    continue
                rows.append(
                    DatasetRow(
                        hash_md5=hash_md5,
                        family=family,
                        creation_day=creation_day,
                        source_csv=str(csv_path),
                        original=dict(raw_row),
                    )
                )
    return rows


def int_cell(value: Any) -> int:
    if value is None or value == "":
        return 0
    return int(float(value))


def first_non_empty(values: Iterable[Any]) -> str:
    for value in values:
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def load_balance_sheet(workbook_path: Path, sheet_name: str) -> BalanceSpec:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"El workbook no tiene la hoja de balance: {sheet_name}")
    ws = workbook[sheet_name]
    title = first_non_empty(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows = list(ws.iter_rows(values_only=True))

    header_index = None
    for index, row in enumerate(rows):
        if row and row[0] == "Familia":
            header_index = index
            break
    if header_index is None:
        raise SystemExit(f"No se encontro encabezado de familias en {sheet_name}")

    headers = [str(value).strip() if value is not None else "" for value in rows[header_index]]
    header_map: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header and header not in header_map:
            header_map[header] = idx
    required = [
        "Familia",
        "Elegibles usados",
        "Train 70% antiguo",
        "Validacion 15% intermedio",
        "Test 15% reciente",
    ]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise SystemExit(f"Faltan columnas en {sheet_name}: {', '.join(missing)}")

    families: dict[str, FamilySplitCounts] = {}
    for row in rows[header_index + 1 :]:
        family = str(row[header_map["Familia"]] or "").strip()
        if not family:
            continue
        if family.lower() == "total":
            break
        counts = FamilySplitCounts(
            family=safe_name(family),
            total=int_cell(row[header_map["Elegibles usados"]]),
            train=int_cell(row[header_map["Train 70% antiguo"]]),
            validation=int_cell(row[header_map["Validacion 15% intermedio"]]),
            test=int_cell(row[header_map["Test 15% reciente"]]),
        )
        families[counts.family] = counts

    return BalanceSpec(sheet=sheet_name, name=title or sheet_name, families=families)


def load_balances(config: dict[str, Any]) -> list[BalanceSpec]:
    workbook_path = Path(str(config["split_workbook"]))
    if not workbook_path.exists():
        raise SystemExit(f"No existe split_workbook: {workbook_path}")
    return [load_balance_sheet(workbook_path, sheet) for sheet in config.get("balances") or []]


def parse_algorithms(config: dict[str, Any]) -> list[AlgorithmSpec]:
    algorithms: list[AlgorithmSpec] = []
    for item in config.get("algorithms") or []:
        if isinstance(item, str):
            algorithms.append(AlgorithmSpec(name=safe_name(item), source=item))
        else:
            name = safe_name(item.get("name") or item.get("source"))
            source = str(item.get("source") or item.get("name") or "").strip()
            if source:
                algorithms.append(AlgorithmSpec(name=name, source=source))
    if not algorithms:
        raise SystemExit("No hay algoritmos configurados.")
    return algorithms


def select_temporal_uniform(rows: list[DatasetRow], target: int) -> list[DatasetRow]:
    if target <= 0:
        return []
    if target >= len(rows):
        return list(rows)
    step = len(rows) / target
    selected = [rows[int((index + 0.5) * step)] for index in range(target)]
    return sorted(selected, key=lambda row: (row.creation_day, row.hash_md5))


def build_assignments(rows: list[DatasetRow], balances: list[BalanceSpec]) -> list[Assignment]:
    by_family: dict[str, list[DatasetRow]] = defaultdict(list)
    for row in rows:
        by_family[row.family].append(row)
    for family_rows in by_family.values():
        family_rows.sort(key=lambda row: (row.creation_day, row.hash_md5))

    assignments: list[Assignment] = []
    for balance in balances:
        for family, counts in sorted(balance.families.items()):
            family_rows = by_family.get(family, [])
            selected = select_temporal_uniform(family_rows, min(counts.total, len(family_rows)))
            train_end = min(counts.train, len(selected))
            validation_end = min(train_end + counts.validation, len(selected))

            for row in selected[:train_end]:
                assignments.append(Assignment(balance=balance, row=row, split="train"))
            for row in selected[train_end:validation_end]:
                assignments.append(Assignment(balance=balance, row=row, split="validacion"))
            for row in selected[validation_end : validation_end + counts.test]:
                assignments.append(Assignment(balance=balance, row=row, split="test"))
    return assignments


def worker_count(config: dict[str, Any]) -> int:
    configured = int(config.get("copy_workers") or 0)
    if configured > 0:
        return configured
    return 32


def source_collections(config: dict[str, Any]) -> list[Path]:
    root = Path(str(config["source_images_root"]))
    configured = [str(value) for value in config.get("source_collections") or []]
    if not configured or configured == ["*"]:
        return sorted(path for path in root.iterdir() if path.is_dir())
    return [root / collection for collection in configured]


def find_image_path(
    hash_md5: str,
    algorithm: AlgorithmSpec,
    collections: list[Path],
    extensions: set[str],
    cache: dict[tuple[str, str], Path | None],
) -> Path | None:
    key = (hash_md5, algorithm.source)
    if key in cache:
        return cache[key]

    for collection in collections:
        algorithm_dir = collection / hash_md5 / "images" / algorithm.source
        if not algorithm_dir.exists():
            continue
        candidates = sorted(
            path
            for path in algorithm_dir.iterdir()
            if path.is_file() and path.suffix.lower() in extensions
        )
        if candidates:
            cache[key] = candidates[0]
            return candidates[0]

    cache[key] = None
    return None


def output_dataset_folder(
    config: dict[str, Any],
    balance: BalanceSpec,
    balance_index: int,
    algorithm: AlgorithmSpec,
    algorithm_index: int,
) -> Path:
    template = str(config.get("folder_template") or DEFAULT_CONFIG["folder_template"])
    folder = template.format(
        dataset_name=safe_name(config.get("dataset_name")),
        balance_index=balance_index,
        algorithm_index=algorithm_index,
        balance_slug=safe_name(balance.name.replace(" - split temporal por familia", "")),
        balance_sheet=safe_name(balance.sheet),
        algorithm=algorithm.name,
        algorithm_source=safe_name(algorithm.source),
    )
    return Path(str(config["output_dir"])) / folder


def destination_path(
    config: dict[str, Any],
    assignment: Assignment,
    balance_index: int,
    algorithm: AlgorithmSpec,
    algorithm_index: int,
    suffix: str,
) -> Path:
    split_names = dict(config.get("split_folder_names") or {})
    split_folder = safe_name(split_names.get(assignment.split) or assignment.split)
    dataset_folder = output_dataset_folder(config, assignment.balance, balance_index, algorithm, algorithm_index)
    filename = f"{assignment.row.hash_md5}{suffix.lower()}"
    return dataset_folder / split_folder / assignment.row.family / filename


def copy_one(source: Path | None, destination: Path, dry_run: bool, overwrite: bool) -> str:
    if source is None:
        return "imagen_no_encontrada"
    if dry_run:
        return "dry_run"
    if destination.exists() and not overwrite:
        return "ya_existia"
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return "copiado"


def write_manifest(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_summary(path: Path, summary: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def expected_counts(assignments: list[Assignment]) -> Counter[tuple[str, str, str]]:
    counter: Counter[tuple[str, str, str]] = Counter()
    for assignment in assignments:
        counter[(assignment.balance.sheet, assignment.split, assignment.row.family)] += 1
    return counter


def process_dataset(config: dict[str, Any], dry_run: bool) -> dict[str, Any]:
    rows = load_dataset_rows(config)
    balances = load_balances(config)
    algorithms = parse_algorithms(config)
    assignments = build_assignments(rows, balances)

    collections = source_collections(config)
    extensions = {str(ext).lower() for ext in config.get("image_extensions") or [".png"]}
    overwrite = bool(config.get("overwrite", False))
    image_cache: dict[tuple[str, str], Path | None] = {}
    manifest_rows: list[dict[str, Any]] = []
    tasks: list[tuple[Path | None, Path, dict[str, Any]]] = []

    balance_index = {balance.sheet: index for index, balance in enumerate(balances, start=1)}
    algorithm_index = {algorithm.name: index for index, algorithm in enumerate(algorithms, start=1)}

    for assignment in assignments:
        for algorithm in algorithms:
            source = find_image_path(
                assignment.row.hash_md5,
                algorithm,
                collections,
                extensions,
                image_cache,
            )
            if source is None:
                suffix = str((config.get("image_extensions") or [".png"])[0])
            else:
                suffix = source.suffix
            destination = destination_path(
                config,
                assignment,
                balance_index[assignment.balance.sheet],
                algorithm,
                algorithm_index[algorithm.name],
                suffix,
            )
            row = {
                "balance_sheet": assignment.balance.sheet,
                "balance_name": assignment.balance.name,
                "algorithm": algorithm.name,
                "algorithm_source": algorithm.source,
                "split": assignment.split,
                "family": assignment.row.family,
                "hash_md5": assignment.row.hash_md5,
                "creation_day": assignment.row.creation_day.isoformat(),
                "source_image": str(source or ""),
                "destination": str(destination),
                "status": "",
            }
            tasks.append((source, destination, row))

    def run_task(task: tuple[Path | None, Path, dict[str, Any]]) -> dict[str, Any]:
        source, destination, row = task
        row = dict(row)
        row["status"] = copy_one(source, destination, dry_run, overwrite)
        return row

    if dry_run or worker_count(config) == 1:
        manifest_rows = [run_task(task) for task in tasks]
    else:
        with ThreadPoolExecutor(max_workers=worker_count(config)) as executor:
            manifest_rows = list(executor.map(run_task, tasks))

    output_dir = Path(str(config["output_dir"]))
    manifest_path = output_dir / "manifest_imagenes.csv"
    summary_path = output_dir / "resumen_imagenes.json"

    status_counts = Counter(row["status"] for row in manifest_rows)
    copied_counts = Counter((row["balance_sheet"], row["algorithm"], row["split"], row["family"], row["status"]) for row in manifest_rows)
    summary: dict[str, Any] = {
        "dataset_name": config.get("dataset_name"),
        "dry_run": dry_run,
        "dataset_rows_temporalmente_elegibles": len(rows),
        "balances": [balance.sheet for balance in balances],
        "algorithms": [{"name": algorithm.name, "source": algorithm.source} for algorithm in algorithms],
        "assignments": len(assignments),
        "image_operations": len(manifest_rows),
        "status_counts": dict(status_counts),
        "expected_counts_by_balance_split_family": {
            "|".join(key): value for key, value in expected_counts(assignments).items()
        },
        "operation_counts": {
            "|".join(key): value for key, value in copied_counts.items()
        },
        "manifest_path": str(manifest_path),
    }

    write_manifest(manifest_path, manifest_rows)
    write_summary(summary_path, summary)
    return summary


def main() -> None:
    args = parse_args()
    config = apply_cli_overrides(load_config(args.config), args)
    summary = process_dataset(config, args.dry_run)

    print("Separacion de dataset de imagenes terminada")
    print(f"Dataset: {summary['dataset_name']}")
    print(f"Dry run: {summary['dry_run']}")
    print(f"Filas temporales elegibles: {summary['dataset_rows_temporalmente_elegibles']:,}")
    print(f"Asignaciones hash/split/balance: {summary['assignments']:,}")
    print(f"Operaciones imagen-algoritmo: {summary['image_operations']:,}")
    for status, count in sorted(summary["status_counts"].items()):
        print(f"{status}: {count:,}")
    print(f"Manifest imagenes: {Path(summary['manifest_path']).resolve()}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Extraer Dataset Final
=====================

Selecciona muestras desde Dataset_V1 o desde la base SQLite generada por
analisis_de_reportes.py, copia los reportes JSON finales y genera un manifest
reproducible.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sqlite3
from collections import Counter, defaultdict, deque
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill


DEFAULT_CONFIG_PATH = "configs_dataset/config_dataset_final.json"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

MANIFEST_FIELDS = [
    "hash_md5",
    "sha1",
    "sha256",
    "lote_origen",
    "familia_probable",
    "familia_confianza",
    "tipo_probable",
    "tipo_confianza",
    "detection_percent",
    "vt_positives",
    "vt_total",
    "fecha_escaneo_vt",
    "dia_escaneo_vt",
    "fecha_creacion_archivo",
    "dia_creacion_archivo",
    "fecha_agregado_virusshare",
    "dia_agregado_virusshare",
    "estrato_deteccion",
    "estrato_fecha",
    "batch_id",
    "fecha_extraccion",
    "config_path",
    "config_hash",
    "criterio_seleccion",
    "reporte_path_original",
    "reporte_path_extraido",
    "estado_copia_reporte",
]

DEFAULT_CONFIG: dict[str, Any] = {
    "db_path": "outputs/Analisis_de_Reportes_Todos.db",
    "source": "dataset_v1",
    "dataset_v1": {
        "csv_dir": "Dataset_V1/csv",
        "csv_files": [],
        "reports_root": "clasificacion",
        "hash_column": "hash_md5",
        "apply_selection": False,
    },
    "output_dir": "outputs/dataset_final",
    "dataset_name": "dataset_final",
    "batch_id": "",
    "copy_reports": False,
    "copy_workers": 0,
    "overwrite_reports": False,
    "organize_reports_by": ["familia_probable", "lote_origen"],
    "filters": {
        "exclude_families": ["sin_inferir"],
        "families": [],
        "family_confidence": ["alta", "media"],
        "types": [],
        "lotes": [],
        "min_detection_percent": 20,
        "max_detection_percent": 100,
        "min_vt_positives": 1,
        "date_field": "dia_escaneo_vt",
        "date_from": "",
        "date_to": "",
        "creation_date_from": "",
        "creation_date_to": "",
    },
    "selection": {
        "min_per_family": 0,
        "max_per_family": 200,
        "max_total": 0,
        "exclude_families_below_min": True,
        "balance_by": ["estrato_deteccion", "estrato_fecha", "lote_origen"],
        "max_per_stratum": 0,
        "sort_by": ["dia_escaneo_vt", "detection_percent DESC", "hash_md5"],
    },
    "stratification": {
        "date_granularity": "month",
        "detection_bands": [
            {"name": "20-40", "min": 20, "max": 40},
            {"name": "40-70", "min": 40, "max": 70},
            {"name": "70-100", "min": 70, "max": 100.0001},
        ],
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrae reportes finales y manifest desde Dataset_V1 o la SQLite de Analisis de Reportes."
    )
    parser.add_argument("--config", default=DEFAULT_CONFIG_PATH, help="Ruta al JSON de configuracion.")
    parser.add_argument(
        "--source",
        choices=["sqlite", "dataset_v1"],
        help="Fuente de muestras: SQLite de analisis o CSV de Dataset_V1.",
    )
    parser.add_argument(
        "--dataset-v1-dir",
        help="Directorio con CSV de Dataset_V1. Solo aplica con --source dataset_v1.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Genera manifest sin copiar reportes.")
    parser.add_argument(
        "--append",
        action="store_true",
        help="Agrega hashes nuevos al manifest existente sin duplicar muestras ya seleccionadas.",
    )
    parser.add_argument(
        "--hashes-only",
        action="store_true",
        help="Genera solo un TXT con un hash MD5 por linea.",
    )
    return parser.parse_args()


def deep_update(base: dict[str, Any], override: dict[str, Any]) -> dict[str, Any]:
    result = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_update(result[key], value)
        else:
            result[key] = value
    return result


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise SystemExit(f"No existe el JSON de configuracion: {path}")
    user_config = json.loads(path.read_text(encoding="utf-8-sig"))
    return deep_update(DEFAULT_CONFIG, user_config)


def config_hash(config: dict[str, Any]) -> str:
    payload = json.dumps(config, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def batch_id(config: dict[str, Any]) -> str:
    configured = str(config.get("batch_id") or "").strip()
    if configured:
        return configured
    return dt.datetime.now(dt.timezone.utc).strftime("%Y%m%d_%H%M%S")


def normalize_list(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(item) for item in value if str(item)]
    return [str(value)]


def placeholders(values: Iterable[Any]) -> str:
    return ",".join("?" for _ in values)


def validate_columns(conn: sqlite3.Connection, required: Iterable[str]) -> None:
    columns = {row[1] for row in conn.execute("PRAGMA table_info(samples)")}
    missing = [column for column in required if column not in columns]
    if missing:
        joined = ", ".join(missing)
        raise SystemExit(
            "La base no tiene columnas necesarias. "
            f"Faltan: {joined}. Vuelve a generar la SQLite con analisis_de_reportes.py."
        )


def build_where(config: dict[str, Any]) -> tuple[str, list[Any]]:
    filters = config["filters"]
    clauses = ["1 = 1"]
    params: list[Any] = []

    exclude_families = normalize_list(filters.get("exclude_families"))
    if exclude_families:
        clauses.append(f"familia_probable NOT IN ({placeholders(exclude_families)})")
        params.extend(exclude_families)

    families = normalize_list(filters.get("families"))
    if families:
        clauses.append(f"familia_probable IN ({placeholders(families)})")
        params.extend(families)

    confidences = normalize_list(filters.get("family_confidence"))
    if confidences:
        clauses.append(f"familia_confianza IN ({placeholders(confidences)})")
        params.extend(confidences)

    malware_types = normalize_list(filters.get("types"))
    if malware_types:
        clauses.append(f"tipo_probable IN ({placeholders(malware_types)})")
        params.extend(malware_types)

    lotes = normalize_list(filters.get("lotes"))
    if lotes:
        clauses.append(f"lote_origen IN ({placeholders(lotes)})")
        params.extend(lotes)

    if filters.get("min_detection_percent") is not None:
        clauses.append("detection_percent >= ?")
        params.append(float(filters["min_detection_percent"]))
    if filters.get("max_detection_percent") is not None:
        clauses.append("detection_percent <= ?")
        params.append(float(filters["max_detection_percent"]))
    if filters.get("min_vt_positives") is not None:
        clauses.append("vt_positives >= ?")
        params.append(int(filters["min_vt_positives"]))

    date_field = str(filters.get("date_field") or "dia_escaneo_vt")
    allowed_date_fields = {"dia_escaneo_vt", "dia_creacion_archivo", "dia_agregado_virusshare"}
    if date_field not in allowed_date_fields:
        raise SystemExit(f"date_field invalido: {date_field}")

    if filters.get("date_from"):
        clauses.append(f"{date_field} >= ?")
        params.append(str(filters["date_from"]))
    if filters.get("date_to"):
        clauses.append(f"{date_field} <= ?")
        params.append(str(filters["date_to"]))

    if filters.get("creation_date_from"):
        clauses.append("dia_creacion_archivo >= ?")
        params.append(str(filters["creation_date_from"]))
    if filters.get("creation_date_to"):
        clauses.append("dia_creacion_archivo <= ?")
        params.append(str(filters["creation_date_to"]))

    return " AND ".join(clauses), params


def load_candidates(config: dict[str, Any]) -> list[dict[str, Any]]:
    source = str(config.get("source") or "sqlite").lower()
    if source == "dataset_v1":
        return load_candidates_from_dataset_v1(config)
    if source != "sqlite":
        raise SystemExit(f"source invalido: {source}. Usa 'sqlite' o 'dataset_v1'.")
    return load_candidates_from_sqlite(config)


def load_candidates_from_sqlite(config: dict[str, Any]) -> list[dict[str, Any]]:
    db_path = Path(config["db_path"])
    if not db_path.exists():
        raise SystemExit(f"No existe la base SQLite: {db_path}")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    validate_columns(
        conn,
        [
            "hash_md5",
            "lote_origen",
            "familia_probable",
            "familia_confianza",
            "tipo_probable",
            "detection_percent",
            "reporte_path",
        ],
    )
    where_sql, params = build_where(config)
    query = f"""
        SELECT *
        FROM samples
        WHERE {where_sql}
        ORDER BY familia_probable, dia_escaneo_vt, detection_percent DESC, hash_md5
    """
    rows = [dict(row) for row in conn.execute(query, params)]
    conn.close()
    return rows


def dataset_v1_csv_paths(config: dict[str, Any]) -> list[Path]:
    dataset_config = dict(config.get("dataset_v1") or {})
    configured_files = normalize_list(dataset_config.get("csv_files"))
    if configured_files:
        paths = [Path(path) for path in configured_files]
    else:
        csv_dir = Path(str(dataset_config.get("csv_dir") or "Dataset_V1/csv"))
        paths = sorted(csv_dir.glob("*.csv"))

    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise SystemExit("No existen CSV de Dataset_V1: " + ", ".join(missing))
    if not paths:
        csv_dir = Path(str(dataset_config.get("csv_dir") or "Dataset_V1/csv"))
        raise SystemExit(f"No se encontraron CSV de Dataset_V1 en: {csv_dir}")
    return paths


def day_from_value(value: Any) -> str:
    text = str(value or "").strip()
    return text[:10] if len(text) >= 10 else text


def dataset_report_path(row: dict[str, Any], config: dict[str, Any]) -> str:
    reports_root = Path(str((config.get("dataset_v1") or {}).get("reports_root") or "clasificacion"))
    lote = str(row.get("lote_origen") or "").strip()
    hash_md5 = safe_name(row.get("hash_md5"))
    if not lote or not hash_md5:
        return ""
    return str(reports_root / lote / "reportes" / "reporte" / f"{hash_md5}.json")


def normalize_dataset_v1_row(raw_row: dict[str, Any], csv_path: Path, config: dict[str, Any]) -> dict[str, Any]:
    hash_column = str((config.get("dataset_v1") or {}).get("hash_column") or "hash_md5")
    hash_md5 = str(raw_row.get(hash_column) or raw_row.get("hash_md5") or "").strip().lower()
    family = str(raw_row.get("familia_probable") or csv_path.stem).strip().lower()
    scan_date = str(raw_row.get("fecha_escaneo_vt") or "").strip()
    creation_day = str(raw_row.get("dia_creacion_archivo") or "").strip()
    row: dict[str, Any] = {
        "hash_md5": hash_md5,
        "sha1": str(raw_row.get("sha1") or "").strip(),
        "sha256": str(raw_row.get("sha256") or "").strip(),
        "lote_origen": str(raw_row.get("lote_origen") or "").strip(),
        "familia_probable": family,
        "familia_confianza": str(raw_row.get("familia_confianza") or "dataset_v1").strip(),
        "tipo_probable": str(raw_row.get("tipo_probable") or "").strip().lower(),
        "tipo_confianza": str(raw_row.get("tipo_confianza") or "dataset_v1").strip(),
        "detection_percent": str(raw_row.get("detection_percent") or "").strip(),
        "vt_positives": str(raw_row.get("vt_positives") or "").strip(),
        "vt_total": str(raw_row.get("vt_total") or "").strip(),
        "fecha_escaneo_vt": scan_date,
        "dia_escaneo_vt": str(raw_row.get("dia_escaneo_vt") or day_from_value(scan_date)).strip(),
        "fecha_creacion_archivo": str(raw_row.get("fecha_creacion_archivo") or creation_day).strip(),
        "dia_creacion_archivo": creation_day,
        "fecha_agregado_virusshare": str(raw_row.get("fecha_agregado_virusshare") or "").strip(),
        "dia_agregado_virusshare": str(raw_row.get("dia_agregado_virusshare") or "").strip(),
    }
    row["reporte_path"] = dataset_report_path(row, config)
    return row


def numeric_between(value: Any, minimum: Any, maximum: Any) -> bool:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    if minimum is not None and number < float(minimum):
        return False
    if maximum is not None and number > float(maximum):
        return False
    return True


def row_matches_filters(row: dict[str, Any], config: dict[str, Any]) -> bool:
    filters = config["filters"]

    exclude_families = set(normalize_list(filters.get("exclude_families")))
    family = str(row.get("familia_probable") or "")
    if family in exclude_families:
        return False

    families = set(normalize_list(filters.get("families")))
    if families and family not in families:
        return False

    confidences = set(normalize_list(filters.get("family_confidence")))
    confidence = str(row.get("familia_confianza") or "")
    if confidences and confidence and confidence != "dataset_v1" and confidence not in confidences:
        return False

    malware_types = set(normalize_list(filters.get("types")))
    if malware_types and str(row.get("tipo_probable") or "") not in malware_types:
        return False

    lotes = set(normalize_list(filters.get("lotes")))
    if lotes and str(row.get("lote_origen") or "") not in lotes:
        return False

    if not numeric_between(
        row.get("detection_percent"),
        filters.get("min_detection_percent"),
        filters.get("max_detection_percent"),
    ):
        return False

    min_vt_positives = filters.get("min_vt_positives")
    if min_vt_positives is not None and str(row.get("vt_positives") or ""):
        try:
            if int(row.get("vt_positives") or 0) < int(min_vt_positives):
                return False
        except (TypeError, ValueError):
            return False

    date_field = str(filters.get("date_field") or "dia_escaneo_vt")
    date_value = str(row.get(date_field) or "")
    if filters.get("date_from") and date_value < str(filters["date_from"]):
        return False
    if filters.get("date_to") and date_value > str(filters["date_to"]):
        return False

    creation_day = str(row.get("dia_creacion_archivo") or "")
    if filters.get("creation_date_from") and creation_day < str(filters["creation_date_from"]):
        return False
    if filters.get("creation_date_to") and creation_day > str(filters["creation_date_to"]):
        return False

    return True


def load_candidates_from_dataset_v1(config: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen_hashes: set[str] = set()
    for csv_path in dataset_v1_csv_paths(config):
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for raw_row in reader:
                row = normalize_dataset_v1_row(raw_row, csv_path, config)
                hash_md5 = str(row.get("hash_md5") or "")
                if not hash_md5 or hash_md5 in seen_hashes:
                    continue
                seen_hashes.add(hash_md5)
                if row_matches_filters(row, config):
                    rows.append(row)
    return rows


def detection_stratum(value: Any, bands: list[dict[str, Any]]) -> str:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return "sin_deteccion"
    for band in bands:
        minimum = float(band.get("min", 0))
        maximum = float(band.get("max", 100))
        if minimum <= number < maximum:
            return str(band.get("name") or f"{minimum}-{maximum}")
    return "fuera_de_bandas"


def date_stratum(row: dict[str, Any], granularity: str, preferred_field: str) -> str:
    day = str(
        row.get(preferred_field)
        or row.get("dia_escaneo_vt")
        or row.get("dia_creacion_archivo")
        or row.get("dia_agregado_virusshare")
        or ""
    )
    if not day:
        return "sin_fecha"
    if granularity == "year":
        return day[:4]
    if granularity == "day":
        return day[:10]
    return day[:7]


def add_strata(rows: list[dict[str, Any]], config: dict[str, Any]) -> None:
    stratification = config["stratification"]
    filters = config["filters"]
    bands = list(stratification.get("detection_bands") or [])
    granularity = str(stratification.get("date_granularity") or "month")
    preferred_date_field = str(filters.get("date_field") or "dia_escaneo_vt")
    for row in rows:
        row["estrato_deteccion"] = detection_stratum(row.get("detection_percent"), bands)
        row["estrato_fecha"] = date_stratum(row, granularity, preferred_date_field)


def sort_value(row: dict[str, Any], expression: str) -> Any:
    field = expression.replace(" DESC", "").replace(" ASC", "").strip()
    value = row.get(field)
    if field in {"detection_percent", "vt_positives", "vt_total"}:
        try:
            value = float(value or 0)
        except (TypeError, ValueError):
            value = 0
    return value if value is not None else ""


def sort_rows(rows: list[dict[str, Any]], sort_by: list[str]) -> list[dict[str, Any]]:
    sorted_rows = list(rows)
    for expression in reversed(sort_by):
        descending = expression.strip().upper().endswith(" DESC")
        sorted_rows.sort(key=lambda row, expr=expression: sort_value(row, expr), reverse=descending)
    return sorted_rows


def select_round_robin(rows: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    selection = config["selection"]
    sort_by = normalize_list(selection.get("sort_by")) or ["hash_md5"]
    balance_by = normalize_list(selection.get("balance_by"))
    max_per_family = int(selection.get("max_per_family") or 0)
    min_per_family = int(selection.get("min_per_family") or 0)
    max_per_stratum = int(selection.get("max_per_stratum") or 0)
    exclude_below_min = bool(selection.get("exclude_families_below_min", True))

    by_family: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_family[str(row.get("familia_probable") or "sin_inferir")].append(row)

    selected_by_family: dict[str, list[dict[str, Any]]] = {}
    skipped_families: set[str] = set()

    for family, family_rows in sorted(by_family.items()):
        if exclude_below_min and min_per_family and len(family_rows) < min_per_family:
            skipped_families.add(family)
            continue

        strata: dict[tuple[Any, ...], deque[dict[str, Any]]] = defaultdict(deque)
        for row in sort_rows(family_rows, sort_by):
            key = tuple(row.get(field, "") for field in balance_by) if balance_by else ("todos",)
            if max_per_stratum and len(strata[key]) >= max_per_stratum:
                continue
            strata[key].append(row)

        family_selected: list[dict[str, Any]] = []
        keys = deque(sorted(strata))
        while keys and (not max_per_family or len(family_selected) < max_per_family):
            key = keys.popleft()
            bucket = strata[key]
            if bucket:
                family_selected.append(bucket.popleft())
            if bucket:
                keys.append(key)

        selected_by_family[family] = family_selected

    selected = flatten_with_global_limit(selected_by_family, int(selection.get("max_total") or 0))
    for row in selected:
        reason = [
            f"familia={row.get('familia_probable', '')}",
            f"confianza={row.get('familia_confianza', '')}",
            f"deteccion={row.get('detection_percent', '')}",
            f"estrato_deteccion={row.get('estrato_deteccion', '')}",
            f"estrato_fecha={row.get('estrato_fecha', '')}",
            f"lote={row.get('lote_origen', '')}",
        ]
        row["criterio_seleccion"] = "; ".join(reason)
    for family in skipped_families:
        pass
    return selected


def select_dataset_v1_rows(rows: list[dict[str, Any]], config: dict[str, Any]) -> list[dict[str, Any]]:
    if bool((config.get("dataset_v1") or {}).get("apply_selection", False)):
        return select_round_robin(rows, config)

    sort_by = normalize_list(config["selection"].get("sort_by")) or ["hash_md5"]
    selected = sort_rows(rows, sort_by)
    for row in selected:
        row["criterio_seleccion"] = "; ".join(
            [
                "fuente=dataset_v1",
                f"familia={row.get('familia_probable', '')}",
                f"deteccion={row.get('detection_percent', '')}",
                f"estrato_deteccion={row.get('estrato_deteccion', '')}",
                f"estrato_fecha={row.get('estrato_fecha', '')}",
                f"lote={row.get('lote_origen', '')}",
            ]
        )
    return selected


def flatten_with_global_limit(selected_by_family: dict[str, list[dict[str, Any]]], max_total: int) -> list[dict[str, Any]]:
    if not max_total:
        return [row for family in sorted(selected_by_family) for row in selected_by_family[family]]

    buckets = deque(
        (family, deque(rows))
        for family, rows in sorted(selected_by_family.items())
        if rows
    )
    selected: list[dict[str, Any]] = []
    while buckets and len(selected) < max_total:
        family, rows = buckets.popleft()
        selected.append(rows.popleft())
        if rows:
            buckets.append((family, rows))
    return selected


def safe_name(value: Any) -> str:
    text = str(value or "sin_valor").strip().lower()
    text = re.sub(r"[^a-z0-9._-]+", "_", text)
    return text.strip("._-") or "sin_valor"


def report_destination(row: dict[str, Any], config: dict[str, Any]) -> Path:
    output_dir = Path(config["output_dir"])
    report_root = output_dir / "reportes"
    parts = [safe_name(row.get(field)) for field in normalize_list(config.get("organize_reports_by"))]
    filename = f"{safe_name(row.get('hash_md5'))}.json"
    return report_root.joinpath(*parts, filename)


def copy_worker_count(requested: int) -> int:
    if requested and requested > 0:
        return requested
    return max(1, min(32, (os.cpu_count() or 4) * 2))


def copy_one_report(row: dict[str, Any], config: dict[str, Any], dry_run: bool) -> tuple[str, str, str, str]:
    source = Path(str(row.get("reporte_path") or row.get("reporte_path_original") or ""))
    destination = report_destination(row, config)

    if dry_run:
        return str(row.get("hash_md5", "")), str(source), str(destination), "dry_run"
    if not bool(config.get("copy_reports", False)):
        return str(row.get("hash_md5", "")), str(source), str(destination), "no_copiado_por_configuracion"
    if not source.exists():
        return str(row.get("hash_md5", "")), str(source), str(destination), "origen_no_existe"
    if destination.exists() and not bool(config.get("overwrite_reports", False)):
        return str(row.get("hash_md5", "")), str(source), str(destination), "ya_existia"

    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    return str(row.get("hash_md5", "")), str(source), str(destination), "copiado"


def copy_reports(rows: list[dict[str, Any]], config: dict[str, Any], dry_run: bool) -> None:
    if not rows:
        return

    rows_by_hash = {str(row.get("hash_md5", "")): row for row in rows}
    workers = copy_worker_count(int(config.get("copy_workers") or 0))
    if workers == 1:
        results = [copy_one_report(row, config, dry_run) for row in rows]
    else:
        with ThreadPoolExecutor(max_workers=workers) as executor:
            results = list(executor.map(lambda row: copy_one_report(row, config, dry_run), rows))

    for hash_md5, source, destination, status in results:
        row = rows_by_hash[hash_md5]
        row["reporte_path_original"] = source
        row["reporte_path_extraido"] = destination
        row["estado_copia_reporte"] = status


def output_paths(config: dict[str, Any]) -> dict[str, Path]:
    output_dir = Path(config["output_dir"])
    dataset_name = safe_name(config.get("dataset_name") or "dataset_final")
    return {
        "csv": output_dir / f"manifest_{dataset_name}.csv",
        "hashes_txt": output_dir / f"hashes_{dataset_name}.txt",
        "xlsx": output_dir / f"manifest_{dataset_name}.xlsx",
        "db": output_dir / f"seleccion_{dataset_name}.db",
        "config": output_dir / f"config_usada_{dataset_name}.json",
    }


def load_existing_manifest(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [normalize_manifest_dict(row) for row in csv.DictReader(handle)]


def normalize_manifest_dict(row: dict[str, Any]) -> dict[str, Any]:
    return {field: row.get(field, "") for field in MANIFEST_FIELDS}


def annotate_batch(rows: list[dict[str, Any]], config_path: Path, config: dict[str, Any], current_batch_id: str) -> None:
    extraction_time = dt.datetime.now(dt.timezone.utc).isoformat()
    current_config_hash = config_hash(config)
    for row in rows:
        row["batch_id"] = current_batch_id
        row["fecha_extraccion"] = extraction_time
        row["config_path"] = str(config_path)
        row["config_hash"] = current_config_hash


def merge_append(existing_rows: list[dict[str, Any]], new_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], int]:
    seen_hashes = {str(row.get("hash_md5", "")) for row in existing_rows if row.get("hash_md5")}
    rows_to_add = [row for row in new_rows if str(row.get("hash_md5", "")) not in seen_hashes]
    skipped = len(new_rows) - len(rows_to_add)
    return existing_rows + rows_to_add, rows_to_add, skipped


def manifest_row(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("hash_md5", ""),
        row.get("sha1", ""),
        row.get("sha256", ""),
        row.get("lote_origen", ""),
        row.get("familia_probable", ""),
        row.get("familia_confianza", ""),
        row.get("tipo_probable", ""),
        row.get("tipo_confianza", ""),
        row.get("detection_percent", ""),
        row.get("vt_positives", ""),
        row.get("vt_total", ""),
        row.get("fecha_escaneo_vt", ""),
        row.get("dia_escaneo_vt", ""),
        row.get("fecha_creacion_archivo", ""),
        row.get("dia_creacion_archivo", ""),
        row.get("fecha_agregado_virusshare", ""),
        row.get("dia_agregado_virusshare", ""),
        row.get("estrato_deteccion", ""),
        row.get("estrato_fecha", ""),
        row.get("batch_id", ""),
        row.get("fecha_extraccion", ""),
        row.get("config_path", ""),
        row.get("config_hash", ""),
        row.get("criterio_seleccion", ""),
        row.get("reporte_path_original", ""),
        row.get("reporte_path_extraido", ""),
        row.get("estado_copia_reporte", ""),
    ]


def styled_header(ws, values: Iterable[Any]) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cells.append(cell)
    return cells


def write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(MANIFEST_FIELDS)
        for row in rows:
            writer.writerow(manifest_row(row))


def write_hashes_txt(rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    hashes = [str(row.get("hash_md5") or "").strip().lower() for row in rows]
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for hash_md5 in hashes:
            if hash_md5:
                handle.write(f"{hash_md5}\n")


def count_by(rows: list[dict[str, Any]], *fields: str) -> list[tuple[Any, ...]]:
    counter: Counter[tuple[Any, ...]] = Counter(tuple(row.get(field, "") for field in fields) for row in rows)
    return [(*key, count) for key, count in sorted(counter.items(), key=lambda item: item[0])]


def append_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[tuple[Any, ...]]) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append(styled_header(ws, headers))
    for row in rows:
        ws.append(list(row))


def write_excel(rows: list[dict[str, Any]], path: Path, config: dict[str, Any], stats: dict[str, int | str]) -> None:
    wb = Workbook(write_only=True)
    append_table_sheet(
        wb,
        "Resumen",
        ["campo", "valor"],
        [
            ("dataset_name", config.get("dataset_name", "")),
            ("modo", stats.get("modo", "")),
            ("candidatas_ejecucion", stats.get("candidatas_ejecucion", 0)),
            ("seleccionadas_ejecucion", stats.get("seleccionadas_ejecucion", 0)),
            ("agregadas_ejecucion", stats.get("agregadas_ejecucion", 0)),
            ("duplicadas_omitidas", stats.get("duplicadas_omitidas", 0)),
            ("total_manifest", len(rows)),
            ("source", config.get("source", "")),
            ("db_path", config.get("db_path", "")),
            ("dataset_v1_csv_dir", (config.get("dataset_v1") or {}).get("csv_dir", "")),
            ("output_dir", config.get("output_dir", "")),
            ("copy_reports", str(config.get("copy_reports", False))),
            ("copy_workers", str(copy_worker_count(int(config.get("copy_workers") or 0)))),
        ],
    )
    append_table_sheet(wb, "Manifest", MANIFEST_FIELDS, [tuple(manifest_row(row)) for row in rows])
    append_table_sheet(wb, "Por_Familia", ["familia_probable", "muestras"], count_by(rows, "familia_probable"))
    append_table_sheet(wb, "Por_Lote", ["lote_origen", "muestras"], count_by(rows, "lote_origen"))
    append_table_sheet(
        wb,
        "Por_Deteccion",
        ["estrato_deteccion", "muestras"],
        count_by(rows, "estrato_deteccion"),
    )
    append_table_sheet(
        wb,
        "Por_Familia_Lote",
        ["familia_probable", "lote_origen", "muestras"],
        count_by(rows, "familia_probable", "lote_origen"),
    )
    wb.save(path)


def write_sqlite(rows: list[dict[str, Any]], path: Path, config: dict[str, Any], stats: dict[str, int | str]) -> None:
    if path.exists():
        path.unlink()
    conn = sqlite3.connect(path)
    conn.execute(
        f"CREATE TABLE manifest ({','.join(f'{field} TEXT' for field in MANIFEST_FIELDS)})"
    )
    conn.executemany(
        f"INSERT INTO manifest ({','.join(MANIFEST_FIELDS)}) VALUES ({','.join('?' for _ in MANIFEST_FIELDS)})",
        [manifest_row(row) for row in rows],
    )
    conn.execute("CREATE TABLE resumen (campo TEXT, valor TEXT)")
    conn.executemany(
        "INSERT INTO resumen VALUES (?, ?)",
        [
            ("dataset_name", str(config.get("dataset_name", ""))),
            ("modo", str(stats.get("modo", ""))),
            ("candidatas_ejecucion", str(stats.get("candidatas_ejecucion", 0))),
            ("seleccionadas_ejecucion", str(stats.get("seleccionadas_ejecucion", 0))),
            ("agregadas_ejecucion", str(stats.get("agregadas_ejecucion", 0))),
            ("duplicadas_omitidas", str(stats.get("duplicadas_omitidas", 0))),
            ("total_manifest", str(len(rows))),
            ("source", str(config.get("source", ""))),
            ("db_path", str(config.get("db_path", ""))),
            ("dataset_v1_csv_dir", str((config.get("dataset_v1") or {}).get("csv_dir", ""))),
            ("output_dir", str(config.get("output_dir", ""))),
            ("copy_workers", str(copy_worker_count(int(config.get("copy_workers") or 0)))),
        ],
    )
    conn.executescript(
        """
        CREATE INDEX idx_manifest_hash ON manifest(hash_md5);
        CREATE INDEX idx_manifest_family ON manifest(familia_probable);
        CREATE INDEX idx_manifest_lote ON manifest(lote_origen);
        """
    )
    conn.commit()
    conn.close()


def write_outputs(rows: list[dict[str, Any]], config: dict[str, Any], stats: dict[str, int | str]) -> dict[str, Path]:
    output_dir = Path(config["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = output_paths(config)
    write_hashes_txt(rows, paths["hashes_txt"])
    write_csv(rows, paths["csv"])
    write_excel(rows, paths["xlsx"], config, stats)
    write_sqlite(rows, paths["db"], config, stats)
    paths["config"].write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
    return paths


def main() -> None:
    args = parse_args()
    config_path = Path(args.config)
    config = load_config(config_path)
    if args.source:
        config["source"] = args.source
    if args.dataset_v1_dir:
        config.setdefault("dataset_v1", {})
        config["dataset_v1"]["csv_dir"] = args.dataset_v1_dir

    candidates = load_candidates(config)
    add_strata(candidates, config)
    if str(config.get("source") or "sqlite").lower() == "dataset_v1":
        selected_current = select_dataset_v1_rows(candidates, config)
    else:
        selected_current = select_round_robin(candidates, config)
    current_batch_id = batch_id(config)
    annotate_batch(selected_current, config_path, config, current_batch_id)

    paths = output_paths(config)
    if args.hashes_only:
        write_hashes_txt(selected_current, paths["hashes_txt"])
        print("Extraccion de hashes terminada")
        print(f"Fuente: {config.get('source', '')}")
        print(f"Hashes escritos: {len(selected_current):,}")
        print(f"TXT hashes: {paths['hashes_txt'].resolve()}")
        return

    existing_rows = load_existing_manifest(paths["csv"]) if args.append else []
    rows_to_write, rows_to_copy, duplicated = (
        merge_append(existing_rows, selected_current) if args.append else (selected_current, selected_current, 0)
    )
    copy_reports(rows_to_copy, config, args.dry_run)

    stats: dict[str, int | str] = {
        "modo": "append" if args.append else "replace",
        "candidatas_ejecucion": len(candidates),
        "seleccionadas_ejecucion": len(selected_current),
        "agregadas_ejecucion": len(rows_to_copy),
        "duplicadas_omitidas": duplicated,
    }
    paths = write_outputs(rows_to_write, config, stats)

    print("Extraccion de dataset final terminada")
    print(f"Modo: {stats['modo']}")
    print(f"Fuente: {config.get('source', '')}")
    print(f"Batch ID: {current_batch_id}")
    print(f"Candidatas en esta ejecucion: {len(candidates):,}")
    print(f"Seleccionadas en esta ejecucion: {len(selected_current):,}")
    print(f"Agregadas al manifest: {len(rows_to_copy):,}")
    if args.append:
        print(f"Duplicadas omitidas: {duplicated:,}")
        print(f"Total acumulado en manifest: {len(rows_to_write):,}")
    print(f"Manifest CSV: {paths['csv'].resolve()}")
    print(f"TXT hashes: {paths['hashes_txt'].resolve()}")
    print(f"Manifest Excel: {paths['xlsx'].resolve()}")
    print(f"SQLite seleccion: {paths['db'].resolve()}")
    print(f"Config usada: {paths['config'].resolve()}")
    if config.get("copy_reports", False) and not args.dry_run:
        print(f"Copy workers utilizados: {copy_worker_count(int(config.get('copy_workers') or 0)):,}")
        copied = sum(1 for row in rows_to_copy if row.get("estado_copia_reporte") == "copiado")
        existing = sum(1 for row in rows_to_copy if row.get("estado_copia_reporte") == "ya_existia")
        missing = sum(1 for row in rows_to_copy if row.get("estado_copia_reporte") == "origen_no_existe")
        print(f"Reportes copiados: {copied:,}")
        print(f"Reportes ya existentes: {existing:,}")
        print(f"Reportes no encontrados: {missing:,}")


if __name__ == "__main__":
    main()

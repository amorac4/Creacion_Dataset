#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Analisis de Reportes
====================

Procesa reportes JSON de VirusShare/VirusTotal y genera:

1) Un Excel consolidado con muestras y resumenes.
2) Una base SQLite para consultas interactivas.

No consulta APIs. Solo analiza los JSON ya descargados.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill


DEFAULT_REPORTS_DIR = r"clasificacion\VirusShare_00499\reportes\reporte"
DEFAULT_OUTPUT_DIR = "outputs"
DEFAULT_NAME = "Analisis_de_Reportes_VirusShare_00499"
EXCEL_MAX_ROWS = 1_048_576
DETECTIONS_ROWS_PER_SHEET = 1_000_000

LABEL_SPLIT_RE = re.compile(r"[^a-zA-Z0-9]+")
HEX_RE = re.compile(r"^[0-9a-f]{8,}$", re.IGNORECASE)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

ENGINE_WEIGHTS = {
    "Microsoft": 2.2,
    "Kaspersky": 2.0,
    "ESET-NOD32": 2.0,
    "BitDefender": 1.8,
    "Fortinet": 1.8,
    "Sophos": 1.8,
    "DrWeb": 1.7,
    "TrendMicro": 1.7,
    "TrendMicro-HouseCall": 1.7,
    "Avira": 1.6,
    "Avast": 1.5,
    "AVG": 1.5,
    "GData": 1.5,
    "F-Secure": 1.5,
    "McAfee": 1.5,
    "McAfee-GW-Edition": 1.5,
    "Malwarebytes": 1.4,
    "Ikarus": 1.4,
    "AhnLab-V3": 1.4,
    "Paloalto": 1.4,
    "Symantec": 1.4,
    "SentinelOne": 1.3,
    "Rising": 1.2,
    "Yandex": 1.2,
}

TYPE_TOKEN_MAP = {
    "adware": "adware",
    "banker": "banker",
    "banking": "banker",
    "backdoor": "backdoor",
    "bot": "bot",
    "botnet": "bot",
    "coinminer": "miner",
    "crypto": "miner",
    "cryptominer": "miner",
    "downloader": "downloader",
    "dropper": "dropper",
    "exploit": "exploit",
    "hacktool": "hacktool",
    "infostealer": "stealer",
    "keylogger": "keylogger",
    "miner": "miner",
    "phish": "phishing",
    "phishing": "phishing",
    "pua": "pua",
    "pup": "pua",
    "ransom": "ransomware",
    "ransomware": "ransomware",
    "redirector": "redirector",
    "riskware": "riskware",
    "rootkit": "rootkit",
    "spyware": "spyware",
    "stealer": "stealer",
    "trojan": "trojan",
    "virus": "virus",
    "worm": "worm",
}

TYPE_PRIORITY = [
    "ransomware",
    "stealer",
    "banker",
    "keylogger",
    "rootkit",
    "backdoor",
    "bot",
    "worm",
    "virus",
    "downloader",
    "dropper",
    "exploit",
    "miner",
    "phishing",
    "redirector",
    "spyware",
    "adware",
    "hacktool",
    "riskware",
    "pua",
    "trojan",
]

NOISE_TOKENS = {
    "a",
    "aa",
    "ab",
    "agent",
    "agen",
    "android",
    "asm",
    "asmalw",
    "asmalwrg",
    "asmalws",
    "bat",
    "behaveslike",
    "clean",
    "classic",
    "cloud",
    "confidence",
    "cve",
    "detect",
    "detected",
    "doc",
    "docx",
    "eldorado",
    "file",
    "gen",
    "generic",
    "genericdet",
    "generickd",
    "generickdz",
    "genericrx",
    "heur",
    "heuristic",
    "high",
    "html",
    "infected",
    "js",
    "javascript",
    "linux",
    "low",
    "macos",
    "malicious",
    "malware",
    "ml",
    "msil",
    "other",
    "outbreak",
    "packed",
    "pdf",
    "possible",
    "powershell",
    "save",
    "script",
    "score",
    "spam",
    "suspicious",
    "tr",
    "troj",
    "unsafe",
    "variant",
    "vba",
    "vbs",
    "w32",
    "w64",
    "webpage",
    "win",
    "win32",
    "win64",
    "windows",
    "x64",
    "x86",
    "xml",
}

NOISE_TOKENS.update(TYPE_TOKEN_MAP.keys())

NOISE_PREFIXES = (
    "adware",
    "backdoor",
    "downloader",
    "dropper",
    "exploit",
    "generic",
    "heur",
    "malware",
    "phish",
    "ransom",
    "spyware",
    "trojan",
    "virus",
    "worm",
)

SAMPLE_FIELDS = [
    "hash_md5",
    "sha1",
    "sha256",
    "extension",
    "filetype",
    "mimetype",
    "size",
    "fecha_escaneo_vt",
    "dia_escaneo_vt",
    "fecha_agregado_virusshare",
    "dia_agregado_virusshare",
    "fecha_creacion_archivo",
    "dia_creacion_archivo",
    "timestamp_creacion_raw",
    "vt_positives",
    "vt_total",
    "detection_ratio",
    "detection_percent",
    "motores_detectaron",
    "familia_probable",
    "familia_confianza",
    "familia_score",
    "familias_top",
    "tipo_probable",
    "tipo_confianza",
    "tipo_score",
    "tipos_top",
    "detecciones_top",
    "permalink",
    "reporte_path",
]

DETECTION_FIELDS = [
    "hash_md5",
    "engine",
    "result",
    "engine_weight",
    "familia_probable",
    "tipo_probable",
    "fecha_escaneo_vt",
    "dia_escaneo_vt",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analisis de Reportes: genera Excel y SQLite desde reportes VirusShare."
    )
    parser.add_argument("--reports-dir", default=DEFAULT_REPORTS_DIR, help="Carpeta con JSON de reportes.")
    parser.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR, help="Carpeta de salida.")
    parser.add_argument("--name", default=DEFAULT_NAME, help="Nombre base para Excel y SQLite.")
    parser.add_argument("--limit", type=int, default=0, help="Procesa solo N reportes para prueba.")
    parser.add_argument("--min-positives", type=int, default=1, help="Minimo de positivos para incluir muestra.")
    parser.add_argument("--min-ratio", type=float, default=0.0, help="Minimo positivos/total para incluir muestra.")
    parser.add_argument("--date-from", help="Fecha minima de escaneo VT: YYYY-MM-DD.")
    parser.add_argument("--date-to", help="Fecha maxima de escaneo VT: YYYY-MM-DD.")
    parser.add_argument("--family", help="Filtra por familia probable, parcial e insensible a mayusculas.")
    parser.add_argument("--type", dest="malware_type", help="Filtra por tipo probable.")
    parser.add_argument(
        "--include-engine-details",
        action="store_true",
        help="Incluye la tabla cruda de detecciones por motor. Aumenta mucho el tamano.",
    )
    parser.add_argument("--no-excel", action="store_true", help="Solo genera SQLite.")
    parser.add_argument("--no-db", action="store_true", help="Solo genera Excel.")
    return parser.parse_args()


def parse_datetime(value: Any) -> dt.datetime | None:
    if not value:
        return None
    text = str(value).strip()
    for candidate in (text, text.replace("Z", "+00:00")):
        try:
            parsed = dt.datetime.fromisoformat(candidate)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=dt.timezone.utc)
            return parsed.astimezone(dt.timezone.utc)
        except ValueError:
            pass
    try:
        parsed = dt.datetime.strptime(text, "%Y-%m-%d %H:%M:%S")
        return parsed.replace(tzinfo=dt.timezone.utc)
    except ValueError:
        return None


def parse_exif_timestamp(value: Any) -> dt.datetime | None:
    if not value:
        return None
    text = str(value).strip()
    if not text or text.startswith("0000:00:00"):
        return None

    for fmt in ("%Y:%m:%d %H:%M:%S%z", "%Y:%m:%d %H:%M:%S"):
        try:
            parsed = dt.datetime.strptime(text, fmt)
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=dt.timezone.utc)
            return parsed.astimezone(dt.timezone.utc)
        except ValueError:
            pass
    return parse_datetime(text)


def parse_date_filter(value: str | None) -> dt.date | None:
    return dt.date.fromisoformat(value) if value else None


def fmt_datetime(value: dt.datetime | None) -> str:
    return value.isoformat() if value else ""


def fmt_date(value: dt.datetime | None) -> str:
    return value.date().isoformat() if value else ""


def label_tokens(label: str) -> list[str]:
    return [token.lower() for token in LABEL_SPLIT_RE.split(label) if token]


def is_hashish(token: str) -> bool:
    return any(ch.isdigit() for ch in token) or bool(HEX_RE.fullmatch(token))


def is_family_noise(token: str) -> bool:
    return token in NOISE_TOKENS or token.startswith(NOISE_PREFIXES)


def engine_weight(engine: str) -> float:
    return ENGINE_WEIGHTS.get(engine, 1.0)


def confidence(score: float, strong: float, medium: float) -> str:
    if score >= strong:
        return "alta"
    if score >= medium:
        return "media"
    if score > 0:
        return "baja"
    return "sin_inferir"


def iter_report_paths(reports_dir: Path, limit: int = 0) -> Iterable[Path]:
    count = 0
    for path in sorted(reports_dir.glob("*.json")):
        yield path
        count += 1
        if limit and count >= limit:
            break


def extract_detections(report: dict[str, Any]) -> list[dict[str, Any]]:
    scans = ((report.get("virustotal") or {}).get("scans") or {})
    detections: list[dict[str, Any]] = []
    if not isinstance(scans, dict):
        return detections

    for engine, details in scans.items():
        if not isinstance(details, dict):
            continue
        result = details.get("result")
        if details.get("detected") is True and result:
            engine_name = str(engine)
            detections.append(
                {
                    "engine": engine_name,
                    "result": str(result),
                    "engine_weight": engine_weight(engine_name),
                }
            )
    return detections


def infer_type_and_family(detections: list[dict[str, Any]]) -> dict[str, Any]:
    type_votes: Counter[str] = Counter()
    family_votes: Counter[str] = Counter()
    total_weight = sum(float(d["engine_weight"]) for d in detections) or 1.0

    for detection in detections:
        tokens = label_tokens(detection["result"])
        weight = float(detection["engine_weight"])

        label_types = {TYPE_TOKEN_MAP[token] for token in tokens if token in TYPE_TOKEN_MAP}
        type_votes.update({malware_type: weight for malware_type in label_types})

        families_in_label: set[str] = set()
        for token in tokens:
            if len(token) < 4:
                continue
            if is_family_noise(token):
                continue
            if is_hashish(token):
                continue
            families_in_label.add(token)
        family_votes.update({family: weight for family in families_in_label})

    tipo = ""
    if type_votes:
        tipo = sorted(
            type_votes,
            key=lambda key: (-type_votes[key], TYPE_PRIORITY.index(key) if key in TYPE_PRIORITY else 999, key),
        )[0]

    familia = ""
    if family_votes:
        familia = family_votes.most_common(1)[0][0]

    type_score = float(type_votes[tipo] / total_weight) if tipo else 0.0
    family_score = float(family_votes[familia] / total_weight) if familia else 0.0
    return {
        "tipo_probable": tipo or "sin_inferir",
        "tipo_confianza": confidence(type_score, strong=0.35, medium=0.15),
        "tipo_score": round(type_score, 4),
        "tipos_top": ";".join(f"{name}:{round(value, 2)}" for name, value in type_votes.most_common(5)),
        "familia_probable": familia or "sin_inferir",
        "familia_confianza": confidence(family_score, strong=0.20, medium=0.08),
        "familia_score": round(family_score, 4),
        "familias_top": ";".join(f"{name}:{round(value, 2)}" for name, value in family_votes.most_common(5)),
    }


def detection_ratio(positives: Any, total: Any) -> float:
    try:
        positives_int = int(positives or 0)
        total_int = int(total or 0)
    except (TypeError, ValueError):
        return 0.0
    return positives_int / total_int if total_int > 0 else 0.0


def row_from_report(path: Path, report: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    vt = report.get("virustotal") or {}
    exif = report.get("exif") or {}
    detections = extract_detections(report)
    inferred = infer_type_and_family(detections)
    scan_dt = parse_datetime(vt.get("scan_date"))
    added_dt = parse_datetime(report.get("added_timestamp"))
    creation_raw = exif.get("TimeStamp") if isinstance(exif, dict) else ""
    creation_dt = parse_exif_timestamp(creation_raw)
    positives = int(vt.get("positives") or 0)
    total = int(vt.get("total") or 0)
    ratio = detection_ratio(positives, total)

    row = {
        "hash_md5": report.get("md5") or report.get("data_structure", {}).get("hash") or path.stem,
        "sha1": report.get("sha1", ""),
        "sha256": report.get("sha256", ""),
        "extension": report.get("extension", ""),
        "filetype": report.get("filetype", ""),
        "mimetype": report.get("mimetype", ""),
        "size": report.get("size", ""),
        "fecha_escaneo_vt": fmt_datetime(scan_dt),
        "dia_escaneo_vt": fmt_date(scan_dt),
        "fecha_agregado_virusshare": fmt_datetime(added_dt),
        "dia_agregado_virusshare": fmt_date(added_dt),
        "fecha_creacion_archivo": fmt_datetime(creation_dt),
        "dia_creacion_archivo": fmt_date(creation_dt),
        "timestamp_creacion_raw": str(creation_raw or ""),
        "vt_positives": positives,
        "vt_total": total,
        "detection_ratio": round(ratio, 4),
        "detection_percent": round(ratio * 100, 2),
        "motores_detectaron": len(detections),
        "familia_probable": inferred["familia_probable"],
        "familia_confianza": inferred["familia_confianza"],
        "familia_score": inferred["familia_score"],
        "familias_top": inferred["familias_top"],
        "tipo_probable": inferred["tipo_probable"],
        "tipo_confianza": inferred["tipo_confianza"],
        "tipo_score": inferred["tipo_score"],
        "tipos_top": inferred["tipos_top"],
        "detecciones_top": ";".join(str(d["result"]) for d in detections[:10]),
        "permalink": vt.get("permalink", ""),
        "reporte_path": str(path),
    }
    return row, detections


def passes_filters(row: dict[str, Any], args: argparse.Namespace, date_from: dt.date | None, date_to: dt.date | None) -> bool:
    if int(row["vt_positives"]) < args.min_positives:
        return False
    if float(row["detection_ratio"]) < args.min_ratio:
        return False

    scan_day = dt.date.fromisoformat(row["dia_escaneo_vt"]) if row["dia_escaneo_vt"] else None
    if date_from and (scan_day is None or scan_day < date_from):
        return False
    if date_to and (scan_day is None or scan_day > date_to):
        return False
    if args.family and args.family.lower() not in str(row["familia_probable"]).lower():
        return False
    if args.malware_type and args.malware_type.lower() not in str(row["tipo_probable"]).lower():
        return False
    return True


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DROP TABLE IF EXISTS samples;
        DROP TABLE IF EXISTS detections;
        DROP TABLE IF EXISTS family_day_counts;
        DROP TABLE IF EXISTS type_day_counts;
        DROP TABLE IF EXISTS family_type_counts;
        DROP TABLE IF EXISTS family_type_day_counts;
        DROP TABLE IF EXISTS processing_errors;

        CREATE TABLE samples (
            hash_md5 TEXT PRIMARY KEY,
            sha1 TEXT,
            sha256 TEXT,
            extension TEXT,
            filetype TEXT,
            mimetype TEXT,
            size INTEGER,
            fecha_escaneo_vt TEXT,
            dia_escaneo_vt TEXT,
            fecha_agregado_virusshare TEXT,
            dia_agregado_virusshare TEXT,
            fecha_creacion_archivo TEXT,
            dia_creacion_archivo TEXT,
            timestamp_creacion_raw TEXT,
            vt_positives INTEGER,
            vt_total INTEGER,
            detection_ratio REAL,
            detection_percent REAL,
            motores_detectaron INTEGER,
            familia_probable TEXT,
            familia_confianza TEXT,
            familia_score REAL,
            familias_top TEXT,
            tipo_probable TEXT,
            tipo_confianza TEXT,
            tipo_score REAL,
            tipos_top TEXT,
            detecciones_top TEXT,
            permalink TEXT,
            reporte_path TEXT
        );

        CREATE TABLE detections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hash_md5 TEXT,
            engine TEXT,
            result TEXT,
            engine_weight REAL,
            familia_probable TEXT,
            tipo_probable TEXT,
            fecha_escaneo_vt TEXT,
            dia_escaneo_vt TEXT
        );

        CREATE TABLE family_day_counts (dia_escaneo_vt TEXT, familia_probable TEXT, muestras INTEGER);
        CREATE TABLE type_day_counts (dia_escaneo_vt TEXT, tipo_probable TEXT, muestras INTEGER);
        CREATE TABLE family_type_counts (familia_probable TEXT, tipo_probable TEXT, muestras INTEGER);
        CREATE TABLE family_type_day_counts (dia_escaneo_vt TEXT, familia_probable TEXT, tipo_probable TEXT, muestras INTEGER);
        CREATE TABLE processing_errors (reporte_path TEXT, error TEXT);
        """
    )


def insert_batch(conn: sqlite3.Connection, samples: list[dict[str, Any]], detections: list[dict[str, Any]]) -> None:
    if samples:
        conn.executemany(
            f"INSERT OR REPLACE INTO samples ({','.join(SAMPLE_FIELDS)}) VALUES ({','.join('?' for _ in SAMPLE_FIELDS)})",
            [[sample[field] for field in SAMPLE_FIELDS] for sample in samples],
        )
    if detections:
        conn.executemany(
            f"INSERT INTO detections ({','.join(DETECTION_FIELDS)}) VALUES ({','.join('?' for _ in DETECTION_FIELDS)})",
            [[detection[field] for field in DETECTION_FIELDS] for detection in detections],
        )


def finalize_database(
    conn: sqlite3.Connection,
    family_day: Counter[tuple[str, str]],
    type_day: Counter[tuple[str, str]],
    family_type: Counter[tuple[str, str]],
    family_type_day: Counter[tuple[str, str, str]],
) -> None:
    conn.executemany(
        "INSERT INTO family_day_counts VALUES (?, ?, ?)",
        [(day, family, count) for (day, family), count in family_day.items()],
    )
    conn.executemany(
        "INSERT INTO type_day_counts VALUES (?, ?, ?)",
        [(day, malware_type, count) for (day, malware_type), count in type_day.items()],
    )
    conn.executemany(
        "INSERT INTO family_type_counts VALUES (?, ?, ?)",
        [(family, malware_type, count) for (family, malware_type), count in family_type.items()],
    )
    conn.executemany(
        "INSERT INTO family_type_day_counts VALUES (?, ?, ?, ?)",
        [(day, family, malware_type, count) for (day, family, malware_type), count in family_type_day.items()],
    )
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_samples_day ON samples(dia_escaneo_vt);
        CREATE INDEX IF NOT EXISTS idx_samples_family ON samples(familia_probable);
        CREATE INDEX IF NOT EXISTS idx_samples_type ON samples(tipo_probable);
        CREATE INDEX IF NOT EXISTS idx_detections_hash ON detections(hash_md5);
        CREATE INDEX IF NOT EXISTS idx_family_day ON family_day_counts(dia_escaneo_vt, familia_probable);
        CREATE INDEX IF NOT EXISTS idx_type_day ON type_day_counts(dia_escaneo_vt, tipo_probable);
        """
    )


def process_reports(args: argparse.Namespace, db_path: Path) -> dict[str, int]:
    reports_dir = Path(args.reports_dir)
    if not reports_dir.is_dir():
        raise SystemExit(f"No existe la carpeta de reportes: {reports_dir}")

    date_from = parse_date_filter(args.date_from)
    date_to = parse_date_filter(args.date_to)

    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    create_schema(conn)

    sample_batch: list[dict[str, Any]] = []
    detection_batch: list[dict[str, Any]] = []
    family_day: Counter[tuple[str, str]] = Counter()
    type_day: Counter[tuple[str, str]] = Counter()
    family_type: Counter[tuple[str, str]] = Counter()
    family_type_day: Counter[tuple[str, str, str]] = Counter()
    metrics = defaultdict(int)

    for path in iter_report_paths(reports_dir, args.limit):
        metrics["reportes_leidos"] += 1
        try:
            report = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            conn.execute("INSERT INTO processing_errors VALUES (?, ?)", (str(path), str(exc)))
            metrics["reportes_corruptos"] += 1
            continue

        if "error_details" in report:
            metrics["reportes_error_api"] += 1
            continue

        row, detections = row_from_report(path, report)
        if not passes_filters(row, args, date_from, date_to):
            metrics["reportes_filtrados_fuera"] += 1
            continue

        sample_batch.append(row)
        metrics["muestras"] += 1
        day = row["dia_escaneo_vt"] or "sin_fecha"
        family = row["familia_probable"] or "sin_inferir"
        malware_type = row["tipo_probable"] or "sin_inferir"
        family_day[(day, family)] += 1
        type_day[(day, malware_type)] += 1
        family_type[(family, malware_type)] += 1
        family_type_day[(day, family, malware_type)] += 1

        metrics["detecciones_leidas"] += len(detections)
        if args.include_engine_details:
            for detection in detections:
                detection_batch.append(
                    {
                        "hash_md5": row["hash_md5"],
                        "engine": detection["engine"],
                        "result": detection["result"],
                        "engine_weight": detection["engine_weight"],
                        "familia_probable": family,
                        "tipo_probable": malware_type,
                        "fecha_escaneo_vt": row["fecha_escaneo_vt"],
                        "dia_escaneo_vt": day,
                    }
                )
                metrics["detecciones_guardadas"] += 1

        if len(sample_batch) >= 1000 or len(detection_batch) >= 20_000:
            insert_batch(conn, sample_batch, detection_batch)
            sample_batch.clear()
            detection_batch.clear()
            conn.commit()

    insert_batch(conn, sample_batch, detection_batch)
    finalize_database(conn, family_day, type_day, family_type, family_type_day)
    conn.commit()
    conn.close()
    return dict(metrics)


def styled_row(ws, values: Iterable[Any], is_header: bool = False) -> list[WriteOnlyCell | Any]:
    if not is_header:
        return list(values)
    cells: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cells.append(cell)
    return cells


def set_sheet_basics(ws, widths: dict[str, int]) -> None:
    ws.freeze_panes = "A2"
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def append_query_sheet(
    wb: Workbook,
    conn: sqlite3.Connection,
    sheet_name: str,
    query: str,
    widths: dict[str, int],
    split_large: bool = False,
) -> tuple[int, int]:
    cursor = conn.execute(query)
    headers = [description[0] for description in cursor.description]
    total_rows = 0
    sheet_count = 0
    current_ws = None
    current_rows = 0

    def new_sheet() -> Any:
        nonlocal sheet_count, current_rows
        sheet_count += 1
        title = f"{sheet_name}_{sheet_count}" if split_large else sheet_name
        ws = wb.create_sheet(title=title[:31])
        set_sheet_basics(ws, widths)
        ws.append(styled_row(ws, headers, is_header=True))
        current_rows = 0
        return ws

    current_ws = new_sheet()
    for row in cursor:
        if split_large and current_rows >= DETECTIONS_ROWS_PER_SHEET:
            current_ws = new_sheet()
        current_ws.append(list(row))
        current_rows += 1
        total_rows += 1

    return total_rows, sheet_count


def add_summary_sheet(wb: Workbook, metrics: dict[str, int], db_path: Path, excel_path: Path) -> None:
    ws = wb.create_sheet(title="Resumen", index=0)
    ws.append(styled_row(ws, ["Analisis de Reportes", ""], is_header=True))
    rows = [
        ("Muestras", f"{metrics.get('muestras', 0):,}"),
        ("Detecciones leidas para inferencia", f"{metrics.get('detecciones_leidas', 0):,}"),
        ("Detecciones crudas guardadas", f"{metrics.get('detecciones_guardadas', 0):,}"),
        ("Reportes leidos", f"{metrics.get('reportes_leidos', 0):,}"),
        ("Reportes corruptos", f"{metrics.get('reportes_corruptos', 0):,}"),
        ("Reportes API error omitidos", f"{metrics.get('reportes_error_api', 0):,}"),
        ("SQLite", str(db_path)),
        ("Excel", str(excel_path)),
    ]
    for row in rows:
        ws.append(row)
    ws.append([])
    ws.append(("Consultas utiles", ""))
    ws.append(("Graficas", "Hoja Graficas"))
    ws.append(("Familias por fecha", "Hoja Familia_por_Dia"))
    ws.append(("Tipos por fecha", "Hoja Tipo_por_Dia"))
    ws.append(("Familia vs tipo", "Hoja Familia_por_Tipo"))
    ws.append(("Cruce fecha/familia/tipo", "Hoja Familia_Tipo_Dia"))
    ws.append(("Nota", "Por defecto no se guardan todos los motores para mantener pequeno el archivo."))
    ws.append(("Motores crudos", "Usa --include-engine-details si necesitas la tabla completa por motor."))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90


def write_chart_block(ws, start_row: int, start_col: int, headers: list[str], rows: list[tuple[Any, ...]]) -> tuple[int, int, int]:
    for offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + offset, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
    return start_row, start_col, start_row + len(rows)


def add_charts_sheet(wb: Workbook, conn: sqlite3.Connection) -> None:
    ws = wb.create_sheet(title="Graficas", index=1)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["J"].width = 22
    ws.column_dimensions["K"].width = 12

    top_types = conn.execute(
        "SELECT tipo_probable, COUNT(*) AS muestras FROM samples GROUP BY tipo_probable ORDER BY muestras DESC LIMIT 12"
    ).fetchall()
    start, col, end = write_chart_block(ws, 1, 1, ["tipo_probable", "muestras"], top_types)
    chart = BarChart()
    chart.title = "Top tipos de malware"
    chart.y_axis.title = "Muestras"
    chart.x_axis.title = "Tipo"
    chart.add_data(Reference(ws, min_col=col + 1, min_row=start, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=col, min_row=start + 1, max_row=end))
    chart.height = 8
    chart.width = 14
    ws.add_chart(chart, "A16")

    top_families = conn.execute(
        "SELECT familia_probable, COUNT(*) AS muestras FROM samples GROUP BY familia_probable ORDER BY muestras DESC LIMIT 15"
    ).fetchall()
    start, col, end = write_chart_block(ws, 1, 4, ["familia_probable", "muestras"], top_families)
    chart = BarChart()
    chart.title = "Top familias"
    chart.y_axis.title = "Muestras"
    chart.x_axis.title = "Familia"
    chart.add_data(Reference(ws, min_col=col + 1, min_row=start, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=col, min_row=start + 1, max_row=end))
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "D20")

    day_counts = conn.execute(
        "SELECT dia_escaneo_vt, COUNT(*) AS muestras FROM samples GROUP BY dia_escaneo_vt ORDER BY dia_escaneo_vt"
    ).fetchall()
    start, col, end = write_chart_block(ws, 1, 7, ["dia_escaneo_vt", "muestras"], day_counts)
    chart = LineChart()
    chart.title = "Muestras por dia de escaneo"
    chart.y_axis.title = "Muestras"
    chart.x_axis.title = "Dia"
    chart.add_data(Reference(ws, min_col=col + 1, min_row=start, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=col, min_row=start + 1, max_row=end))
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "G20")

    confidence_counts = conn.execute(
        "SELECT familia_confianza, COUNT(*) AS muestras FROM samples GROUP BY familia_confianza ORDER BY muestras DESC"
    ).fetchall()
    start, col, end = write_chart_block(ws, 1, 10, ["familia_confianza", "muestras"], confidence_counts)
    chart = PieChart()
    chart.title = "Confianza de familia"
    chart.add_data(Reference(ws, min_col=col + 1, min_row=start, max_row=end), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=col, min_row=start + 1, max_row=end))
    chart.height = 8
    chart.width = 12
    ws.add_chart(chart, "J12")


def export_excel(db_path: Path, excel_path: Path, metrics: dict[str, int]) -> None:
    conn = sqlite3.connect(db_path)
    include_engine_details = bool(metrics.get("detecciones_guardadas", 0))
    wb = Workbook(write_only=include_engine_details)
    if not include_engine_details and "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    add_summary_sheet(wb, metrics, db_path, excel_path)
    if not include_engine_details:
        add_charts_sheet(wb, conn)
    append_query_sheet(
        wb,
        conn,
        "Muestras",
        "SELECT * FROM samples ORDER BY dia_escaneo_vt, familia_probable, tipo_probable",
        {"A": 34, "B": 42, "C": 64, "H": 24, "J": 24, "P": 22, "Q": 18, "U": 18, "V": 18, "Y": 70},
    )
    append_query_sheet(
        wb,
        conn,
        "Familia_por_Dia",
        "SELECT dia_escaneo_vt, familia_probable, muestras FROM family_day_counts ORDER BY dia_escaneo_vt, muestras DESC",
        {"A": 18, "B": 24, "C": 12},
    )
    append_query_sheet(
        wb,
        conn,
        "Tipo_por_Dia",
        "SELECT dia_escaneo_vt, tipo_probable, muestras FROM type_day_counts ORDER BY dia_escaneo_vt, muestras DESC",
        {"A": 18, "B": 18, "C": 12},
    )
    append_query_sheet(
        wb,
        conn,
        "Familia_por_Tipo",
        "SELECT familia_probable, tipo_probable, muestras FROM family_type_counts ORDER BY muestras DESC",
        {"A": 24, "B": 18, "C": 12},
    )
    append_query_sheet(
        wb,
        conn,
        "Familia_Tipo_Dia",
        "SELECT dia_escaneo_vt, familia_probable, tipo_probable, muestras FROM family_type_day_counts ORDER BY dia_escaneo_vt, muestras DESC",
        {"A": 18, "B": 24, "C": 18, "D": 12},
    )
    append_query_sheet(
        wb,
        conn,
        "Top_Familias",
        "SELECT familia_probable, COUNT(*) AS muestras FROM samples GROUP BY familia_probable ORDER BY muestras DESC",
        {"A": 24, "B": 12},
    )
    append_query_sheet(
        wb,
        conn,
        "Top_Tipos",
        "SELECT tipo_probable, COUNT(*) AS muestras FROM samples GROUP BY tipo_probable ORDER BY muestras DESC",
        {"A": 18, "B": 12},
    )
    if metrics.get("detecciones_guardadas", 0):
        append_query_sheet(
            wb,
            conn,
            "Detecciones",
            "SELECT hash_md5, engine, result, engine_weight, familia_probable, tipo_probable, fecha_escaneo_vt, dia_escaneo_vt FROM detections ORDER BY hash_md5",
            {"A": 34, "B": 22, "C": 52, "D": 14, "E": 22, "F": 18, "G": 24, "H": 16},
            split_large=True,
        )
    conn.close()
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    db_path = output_dir / f"{args.name}.db"
    excel_path = output_dir / f"{args.name}.xlsx"

    metrics = process_reports(args, db_path)
    if args.no_db:
        # Excel still needs the transient DB for efficient aggregation.
        pass
    if not args.no_excel:
        export_excel(db_path, excel_path, metrics)
    if args.no_db and db_path.exists():
        db_path.unlink()

    print("Analisis de Reportes terminado")
    print(f"Muestras: {metrics.get('muestras', 0):,}")
    print(f"Detecciones leidas para inferencia: {metrics.get('detecciones_leidas', 0):,}")
    print(f"Detecciones crudas guardadas: {metrics.get('detecciones_guardadas', 0):,}")
    if not args.no_excel:
        print(f"Excel: {excel_path.resolve()}")
    if not args.no_db:
        print(f"SQLite: {db_path.resolve()}")


if __name__ == "__main__":
    main()

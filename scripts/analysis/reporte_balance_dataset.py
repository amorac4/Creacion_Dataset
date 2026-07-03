#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Reporte de balance del dataset actual.

Lee la fuente consolidada de analisis (`outputs/Analisis_de_Reportes_Todos.db`
si existe; si no, usa el XLSX equivalente) y los CSV curados en `Dataset_V1/csv`.
Genera tablas CSV y un resumen Markdown para revisar sesgos de familia, fechas y
lotes de origen.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import math
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_DB = Path("outputs/Analisis_de_Reportes_Todos.db")
DEFAULT_XLSX = Path("outputs/Analisis_de_Reportes_Todos.xlsx")
DEFAULT_DATASET_CSV_DIR = Path("Dataset_V1/csv")
DEFAULT_OUTPUT_DIR = Path("outputs/balance_dataset_actual")

REQUIRED_FIELDS = [
    "hash_md5",
    "lote_origen",
    "familia_probable",
    "tipo_probable",
    "detection_percent",
    "fecha_escaneo_vt",
    "dia_creacion_archivo",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera un reporte de balance para la fuente completa y Dataset_V1."
    )
    parser.add_argument("--db", type=Path, default=DEFAULT_DB, help="SQLite consolidada esperada.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="XLSX consolidado de respaldo.")
    parser.add_argument(
        "--dataset-csv-dir",
        type=Path,
        default=DEFAULT_DATASET_CSV_DIR,
        help="Directorio con CSV finales por familia.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directorio de salida.")
    return parser.parse_args()


def norm_text(value: Any, default: str = "sin_dato") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def norm_family(value: Any) -> str:
    return norm_text(value, "sin_inferir").lower()


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_date(value: Any) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    if "T" in text:
        try:
            return dt.datetime.fromisoformat(text).date()
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def year_month(date_value: dt.date | None) -> str:
    return date_value.strftime("%Y-%m") if date_value else "sin_fecha"


def year_value(date_value: dt.date | None) -> str:
    return str(date_value.year) if date_value else "sin_fecha"


def date_status(date_value: dt.date | None, today: dt.date) -> str:
    if date_value is None:
        return "sin_fecha"
    if date_value.year < 1995:
        return "pre_1995"
    if date_value > today:
        return "futura"
    return "plausible"


def load_rows_from_db(path: Path) -> list[dict[str, Any]]:
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    try:
        columns = {row[1] for row in con.execute("PRAGMA table_info(samples)")}
        missing = [field for field in REQUIRED_FIELDS if field not in columns]
        if missing:
            joined = ", ".join(missing)
            raise SystemExit(f"La tabla samples no tiene columnas requeridas: {joined}")
        query = "SELECT * FROM samples"
        return [dict(row) for row in con.execute(query)]
    finally:
        con.close()


def load_rows_from_xlsx(path: Path, sheet_name: str = "Muestras") -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"El XLSX no contiene la hoja {sheet_name!r}: {path}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    missing = [field for field in REQUIRED_FIELDS if field not in header]
    if missing:
        joined = ", ".join(missing)
        raise SystemExit(f"La hoja {sheet_name!r} no tiene columnas requeridas: {joined}")
    return [dict(zip(header, row)) for row in rows]


def load_dataset_csv_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for csv_path in sorted(path.glob("*.csv")):
        with csv_path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                row["csv_origen"] = csv_path.name
                rows.append(row)
    return rows


def counter_to_rows(counter: Counter[tuple[Any, ...] | Any], headers: list[str]) -> list[dict[str, Any]]:
    total = sum(counter.values())
    output: list[dict[str, Any]] = []
    for key, count in counter.most_common():
        values = key if isinstance(key, tuple) else (key,)
        row = {header: value for header, value in zip(headers, values)}
        row["muestras"] = count
        row["porcentaje"] = round((count / total) * 100, 4) if total else 0
        output.append(row)
    return output


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def analyze_rows(rows: list[dict[str, Any]], label: str, today: dt.date) -> dict[str, Any]:
    families: Counter[str] = Counter()
    lotes: Counter[str] = Counter()
    months: Counter[str] = Counter()
    years: Counter[str] = Counter()
    family_month: Counter[tuple[str, str]] = Counter()
    family_lote: Counter[tuple[str, str]] = Counter()
    family_year: Counter[tuple[str, str]] = Counter()
    date_quality: Counter[str] = Counter()
    type_counts: Counter[str] = Counter()
    family_detection: defaultdict[str, list[float]] = defaultdict(list)

    for row in rows:
        family = norm_family(row.get("familia_probable"))
        lote = norm_text(row.get("lote_origen"))
        malware_type = norm_text(row.get("tipo_probable"))
        creation_date = parse_date(row.get("dia_creacion_archivo") or row.get("fecha_creacion_archivo"))
        month = year_month(creation_date)
        year = year_value(creation_date)
        detection = parse_float(row.get("detection_percent"))

        families[family] += 1
        lotes[lote] += 1
        months[month] += 1
        years[year] += 1
        family_month[(family, month)] += 1
        family_lote[(family, lote)] += 1
        family_year[(family, year)] += 1
        date_quality[date_status(creation_date, today)] += 1
        type_counts[malware_type] += 1
        if detection is not None:
            family_detection[family].append(detection)

    total = len(rows)
    dominant_family = families.most_common(1)[0] if families else ("sin_dato", 0)
    dominant_lote = lotes.most_common(1)[0] if lotes else ("sin_dato", 0)
    dominant_month = months.most_common(1)[0] if months else ("sin_fecha", 0)
    hhi_family = sum((count / total) ** 2 for count in families.values()) if total else 0
    hhi_lote = sum((count / total) ** 2 for count in lotes.values()) if total else 0
    dated_total = total - date_quality["sin_fecha"]

    detection_rows = []
    for family, values in family_detection.items():
        values_sorted = sorted(values)
        avg = sum(values_sorted) / len(values_sorted)
        mid = len(values_sorted) // 2
        if len(values_sorted) % 2:
            median = values_sorted[mid]
        else:
            median = (values_sorted[mid - 1] + values_sorted[mid]) / 2
        detection_rows.append(
            {
                "familia_probable": family,
                "muestras_con_detection": len(values_sorted),
                "detection_promedio": round(avg, 4),
                "detection_mediana": round(median, 4),
                "detection_min": round(values_sorted[0], 4),
                "detection_max": round(values_sorted[-1], 4),
            }
        )
    detection_rows.sort(key=lambda row: (-row["muestras_con_detection"], row["familia_probable"]))

    return {
        "label": label,
        "total": total,
        "unique_families": len(families),
        "unique_lotes": len(lotes),
        "dated_total": dated_total,
        "dominant_family": dominant_family,
        "dominant_lote": dominant_lote,
        "dominant_month": dominant_month,
        "hhi_family": hhi_family,
        "hhi_lote": hhi_lote,
        "families": families,
        "lotes": lotes,
        "months": months,
        "years": years,
        "family_month": family_month,
        "family_lote": family_lote,
        "family_year": family_year,
        "date_quality": date_quality,
        "type_counts": type_counts,
        "family_detection_rows": detection_rows,
    }


def write_analysis_tables(output_dir: Path, prefix: str, analysis: dict[str, Any]) -> None:
    write_csv(
        output_dir / f"{prefix}_conteos_por_familia.csv",
        counter_to_rows(analysis["families"], ["familia_probable"]),
    )
    write_csv(
        output_dir / f"{prefix}_conteos_por_lote.csv",
        counter_to_rows(analysis["lotes"], ["lote_origen"]),
    )
    write_csv(
        output_dir / f"{prefix}_conteos_por_anio.csv",
        counter_to_rows(analysis["years"], ["anio_creacion_archivo"]),
    )
    write_csv(
        output_dir / f"{prefix}_conteos_por_mes.csv",
        counter_to_rows(analysis["months"], ["mes_creacion_archivo"]),
    )
    write_csv(
        output_dir / f"{prefix}_familia_por_mes.csv",
        counter_to_rows(analysis["family_month"], ["familia_probable", "mes_creacion_archivo"]),
    )
    write_csv(
        output_dir / f"{prefix}_familia_por_lote.csv",
        counter_to_rows(analysis["family_lote"], ["familia_probable", "lote_origen"]),
    )
    write_csv(
        output_dir / f"{prefix}_familia_por_anio.csv",
        counter_to_rows(analysis["family_year"], ["familia_probable", "anio_creacion_archivo"]),
    )
    write_csv(
        output_dir / f"{prefix}_calidad_fecha_creacion.csv",
        counter_to_rows(analysis["date_quality"], ["estado_fecha_creacion"]),
    )
    write_csv(
        output_dir / f"{prefix}_conteos_por_tipo.csv",
        counter_to_rows(analysis["type_counts"], ["tipo_probable"]),
    )
    write_csv(
        output_dir / f"{prefix}_deteccion_por_familia.csv",
        analysis["family_detection_rows"],
    )


def pct(count: int, total: int) -> float:
    return round((count / total) * 100, 2) if total else 0.0


def top_lines(counter: Counter[Any], total: int, limit: int = 8) -> list[str]:
    lines = []
    for key, count in counter.most_common(limit):
        lines.append(f"- {key}: {count:,} ({pct(count, total)}%)")
    return lines


def month_span(months: Counter[str]) -> str:
    valid = sorted(month for month in months if month != "sin_fecha")
    if not valid:
        return "sin meses validos"
    return f"{valid[0]} a {valid[-1]}"


def build_bias_notes(source: dict[str, Any], dataset: dict[str, Any], source_kind: str) -> list[str]:
    notes: list[str] = []
    total = dataset["total"]
    if source_kind != "db":
        notes.append(
            "La SQLite consolidada esperada no esta presente; el analisis de fuente completa uso el XLSX consolidado como respaldo."
        )

    dominant_family, dominant_family_count = dataset["dominant_family"]
    if total and dominant_family_count / total >= 0.25:
        notes.append(
            f"La familia dominante en Dataset_V1 es {dominant_family} con {dominant_family_count:,} muestras "
            f"({pct(dominant_family_count, total)}%). Esto puede hacer que tendencias temporales parezcan de todo el malware "
            "cuando en realidad reflejan una familia."
        )

    top_two_families = dataset["families"].most_common(2)
    if total and len(top_two_families) == 2:
        top_two_count = sum(count for _, count in top_two_families)
        if top_two_count / total >= 0.50:
            names = " y ".join(family for family, _ in top_two_families)
            notes.append(
                f"Las dos familias principales ({names}) suman {top_two_count:,} muestras "
                f"({pct(top_two_count, total)}%). Cualquier tendencia agregada deberia contrastarse por familia."
            )

    dominant_lote, dominant_lote_count = dataset["dominant_lote"]
    if total and dominant_lote_count / total >= 0.60:
        notes.append(
            f"Dataset_V1 depende fuertemente de {dominant_lote}: {dominant_lote_count:,} muestras "
            f"({pct(dominant_lote_count, total)}%). Un lote de VirusShare no es una muestra temporal aleatoria del ecosistema."
        )

    missing_dates = dataset["date_quality"]["sin_fecha"]
    pre_1995 = dataset["date_quality"]["pre_1995"]
    future = dataset["date_quality"]["futura"]
    questionable = missing_dates + pre_1995 + future
    if total and questionable / total >= 0.05:
        notes.append(
            f"Hay {questionable:,} muestras ({pct(questionable, total)}%) con fecha de creacion faltante o anomala "
            f"(sin_fecha={missing_dates:,}, pre_1995={pre_1995:,}, futura={future:,}). "
            "El campo dia_creacion_archivo debe tratarse como timestamp de archivo/compilacion, no como fecha confiable de aparicion."
        )

    top_month, top_month_count = dataset["dominant_month"]
    if top_month != "sin_fecha" and total and top_month_count / total >= 0.15:
        notes.append(
            f"El mes mas cargado es {top_month} con {top_month_count:,} muestras ({pct(top_month_count, total)}%). "
            "Un pico mensual asi puede dominar modelos o graficas de evolucion temporal."
        )

    family_month_total = sum(dataset["family_month"].values())
    dominant_family_months = [
        (family, month, count)
        for (family, month), count in dataset["family_month"].most_common(8)
        if month != "sin_fecha" and total and count / total >= 0.04
    ]
    if dominant_family_months:
        fragments = [
            f"{family} en {month} ({count:,}; {pct(count, total)}%)"
            for family, month, count in dominant_family_months[:5]
        ]
        notes.append(
            "Hay concentraciones familia-mes muy marcadas: "
            + "; ".join(fragments)
            + ". Esto sugiere que el eje temporal esta parcialmente definido por bloques de seleccion familiar."
        )

    if dataset["unique_families"] <= 10:
        notes.append(
            f"Dataset_V1 contiene solo {dataset['unique_families']} familias. Para una tesis sobre evolucion temporal del malware, "
            "esto apoya mejor afirmaciones sobre estas familias curadas que sobre el ecosistema completo."
        )

    source_total = source["total"]
    if source_total and total:
        coverage = total / source_total
        notes.append(
            f"Dataset_V1 representa {total:,} de {source_total:,} muestras de la fuente consolidada ({coverage * 100:.2f}%). "
            "La seleccion curada mejora control experimental, pero reduce representatividad poblacional."
        )

    return notes


def write_markdown_report(
    output_dir: Path,
    source: dict[str, Any],
    dataset: dict[str, Any],
    source_path: Path,
    source_kind: str,
    db_path: Path,
    xlsx_path: Path,
    dataset_dir: Path,
) -> None:
    notes = build_bias_notes(source, dataset, source_kind)
    md: list[str] = []
    md.append("# Reporte de balance del dataset actual")
    md.append("")
    md.append(f"Generado: {dt.datetime.now().isoformat(timespec='seconds')}")
    md.append("")
    md.append("## Fuentes")
    md.append("")
    md.append(f"- Fuente consolidada usada: `{source_path}` ({source_kind})")
    md.append(f"- SQLite esperada: `{db_path}` ({'existe' if db_path.exists() else 'no encontrada'})")
    md.append(f"- XLSX de respaldo: `{xlsx_path}` ({'existe' if xlsx_path.exists() else 'no encontrado'})")
    md.append(f"- Dataset curado: `{dataset_dir}`")
    md.append("")
    md.append("## Resumen ejecutivo")
    md.append("")
    md.append(f"- Fuente consolidada: {source['total']:,} muestras, {source['unique_families']:,} familias, {source['unique_lotes']:,} lotes.")
    md.append(f"- Dataset_V1: {dataset['total']:,} muestras, {dataset['unique_families']:,} familias, {dataset['unique_lotes']:,} lotes.")
    md.append(f"- Cobertura temporal por `dia_creacion_archivo` en Dataset_V1: {month_span(dataset['months'])}.")
    md.append(f"- Muestras de Dataset_V1 con fecha plausible: {dataset['date_quality']['plausible']:,} ({pct(dataset['date_quality']['plausible'], dataset['total'])}%).")
    md.append("")
    md.append("## Conteos principales de Dataset_V1")
    md.append("")
    md.append("### Familias")
    md.extend(top_lines(dataset["families"], dataset["total"], limit=12))
    md.append("")
    md.append("### Lotes")
    md.extend(top_lines(dataset["lotes"], dataset["total"], limit=12))
    md.append("")
    md.append("### Familias por lote mas frecuentes")
    for (family, lote), count in dataset["family_lote"].most_common(12):
        md.append(f"- {family} / {lote}: {count:,} ({pct(count, dataset['total'])}%)")
    md.append("")
    md.append("### Anios de creacion del archivo")
    for key, count in sorted(dataset["years"].items(), key=lambda item: str(item[0])):
        md.append(f"- {key}: {count:,} ({pct(count, dataset['total'])}%)")
    md.append("")
    md.append("### Meses mas frecuentes")
    md.extend(top_lines(dataset["months"], dataset["total"], limit=15))
    md.append("")
    md.append("### Familia-mes mas frecuentes")
    for (family, month), count in dataset["family_month"].most_common(12):
        md.append(f"- {family} / {month}: {count:,} ({pct(count, dataset['total'])}%)")
    md.append("")
    md.append("## Contexto de la fuente consolidada")
    md.append("")
    md.append("### Familias principales en la fuente")
    md.extend(top_lines(source["families"], source["total"], limit=10))
    md.append("")
    md.append("### Calidad de fecha en la fuente")
    for status, count in source["date_quality"].most_common():
        md.append(f"- {status}: {count:,} ({pct(count, source['total'])}%)")
    md.append("")
    md.append("## Sesgos fuertes para una tesis temporal")
    md.append("")
    for note in notes:
        md.append(f"- {note}")
    md.append("")
    md.append("## Archivos generados")
    md.append("")
    for path in sorted(output_dir.glob("*.csv")):
        md.append(f"- `{path.name}`")
    md.append("")
    md.append("## Lectura recomendada")
    md.append("")
    md.append(
        "Usa `dataset_v1_familia_por_mes.csv` y `dataset_v1_familia_por_anio.csv` para separar evolucion temporal por familia. "
        "Evita interpretar `dia_creacion_archivo` como fecha historica absoluta sin filtrar `dataset_v1_calidad_fecha_creacion.csv`."
    )

    (output_dir / "reporte_balance_dataset_actual.md").write_text("\n".join(md), encoding="utf-8")


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    if args.db.exists():
        source_kind = "db"
        source_path = args.db
        source_rows = load_rows_from_db(args.db)
    elif args.xlsx.exists():
        source_kind = "xlsx_respaldo"
        source_path = args.xlsx
        source_rows = load_rows_from_xlsx(args.xlsx)
    else:
        raise SystemExit(f"No existe {args.db} ni {args.xlsx}; no hay fuente consolidada para analizar.")

    dataset_rows = load_dataset_csv_rows(args.dataset_csv_dir)
    if not dataset_rows:
        raise SystemExit(f"No se encontraron CSV en {args.dataset_csv_dir}")

    today = dt.date.today()
    source_analysis = analyze_rows(source_rows, "fuente_consolidada", today)
    dataset_analysis = analyze_rows(dataset_rows, "dataset_v1", today)

    write_analysis_tables(args.output_dir, "fuente_consolidada", source_analysis)
    write_analysis_tables(args.output_dir, "dataset_v1", dataset_analysis)
    write_markdown_report(
        args.output_dir,
        source_analysis,
        dataset_analysis,
        source_path,
        source_kind,
        args.db,
        args.xlsx,
        args.dataset_csv_dir,
    )

    print(f"Reporte generado en: {args.output_dir}")
    print(f"Fuente consolidada usada: {source_path} ({source_kind})")
    print(f"Filas fuente consolidada: {source_analysis['total']:,}")
    print(f"Filas Dataset_V1: {dataset_analysis['total']:,}")


if __name__ == "__main__":
    main()

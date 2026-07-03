#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Generar Excel por Dataset de Imagenes
=====================================

Construye un workbook XLSX por cada dataset (balance + algoritmo) usando:

- manifest_imagenes.csv generado por separar_dataset_imagenes.py;
- CSV curados de Dataset_V1;
- JSON de referencia por hash generados en Lab_Creacion_Dataset.

Cada workbook incluye resumen, graficas, distribuciones, fechas, metricas y
detalle por muestra.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_MANIFEST = Path("outputs/datasets_imagenes/manifest_imagenes.csv")
DEFAULT_DATASET_CSV_DIR = Path("Dataset_V1/csv")
DEFAULT_JSON_ROOT = Path("C:/Users/ADOLF/Desktop/Repositorios/Lab_Creacion_Dataset/data/DatasetV1/DatasetV1")
DEFAULT_OUTPUT_DIR = Path("outputs/excel_datasets_imagenes")
AVAILABLE_STATUSES = {"copiado", "ya_existia"}
SPLIT_ORDER = {"train": 0, "validacion": 1, "val": 1, "test": 2}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


@dataclass
class ManifestRow:
    dataset_folder: str
    actual_dataset_folder: str
    balance_sheet: str
    balance_name: str
    algorithm: str
    algorithm_source: str
    split: str
    family: str
    hash_md5: str
    creation_day: str
    source_image: str
    destination: str
    actual_destination: str
    status: str
    destination_exists: bool


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera un XLSX por cada dataset de imagenes.")
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST, help="Manifest de imagenes.")
    parser.add_argument("--dataset-csv-dir", type=Path, default=DEFAULT_DATASET_CSV_DIR, help="CSV por familia.")
    parser.add_argument("--json-root", type=Path, default=DEFAULT_JSON_ROOT, help="Raiz con JSON por hash.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directorio de salida.")
    parser.add_argument(
        "--datasets",
        nargs="+",
        help="Datasets/carpetas a generar. Si se omite, genera todos.",
    )
    parser.add_argument("--no-file-check", action="store_true", help="No valida archivos destino en disco.")
    parser.add_argument(
        "--deep-image-json",
        action="store_true",
        help="Lee image_analysis.json completo para dimensiones por algoritmo. Es mucho mas lento.",
    )
    return parser.parse_args()


def safe_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    chars = [char if char.isalnum() or char in "._-" else "_" for char in text]
    cleaned = "".join(chars).strip("._-")
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned or "sin_valor"


def sheet_name(value: str) -> str:
    invalid = "[]:*?/\\"
    cleaned = "".join("_" if char in invalid else char for char in value)
    return cleaned[:31] or "Hoja"


def dataset_folder_from_destination(destination: str) -> str:
    parts = Path(destination).parts
    if len(parts) >= 4:
        return parts[-4]
    return "sin_dataset"


def dataset_code(folder: str) -> str:
    text = str(folder or "")
    text = text.split("__", 1)[0]
    text = text.split("_", 1)[0]
    return text


def actual_folder_map(output_dir: Path) -> dict[str, str]:
    folders: dict[str, str] = {}
    if not output_dir.exists():
        return folders
    for path in sorted(output_dir.iterdir()):
        if path.is_dir():
            folders.setdefault(dataset_code(path.name), path.name)
    return folders


def resolve_actual_destination(path_text: str, manifest_path: Path, folders_by_code: dict[str, str]) -> tuple[str, Path]:
    path = Path(path_text)
    parts = path.parts
    if len(parts) < 4:
        resolved = path if path.is_absolute() else (manifest_path.parent.parent.parent / path)
        return dataset_folder_from_destination(path_text), resolved.resolve()

    manifest_folder = parts[-4]
    actual_folder = folders_by_code.get(dataset_code(manifest_folder), manifest_folder)
    actual_path = manifest_path.parent / actual_folder / parts[-3] / parts[-2] / parts[-1]
    return actual_folder, actual_path.resolve()


def load_manifest(path: Path, check_files: bool) -> list[ManifestRow]:
    if not path.exists():
        raise SystemExit(f"No existe el manifest: {path}")

    rows: list[ManifestRow] = []
    folders_by_code = actual_folder_map(path.parent)
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for raw in reader:
            destination = str(raw.get("destination") or "")
            actual_folder, actual_destination = resolve_actual_destination(destination, path, folders_by_code)
            rows.append(
                ManifestRow(
                    dataset_folder=dataset_folder_from_destination(destination),
                    actual_dataset_folder=actual_folder,
                    balance_sheet=str(raw.get("balance_sheet") or ""),
                    balance_name=str(raw.get("balance_name") or ""),
                    algorithm=str(raw.get("algorithm") or ""),
                    algorithm_source=str(raw.get("algorithm_source") or ""),
                    split=str(raw.get("split") or ""),
                    family=str(raw.get("family") or ""),
                    hash_md5=str(raw.get("hash_md5") or "").lower(),
                    creation_day=str(raw.get("creation_day") or ""),
                    source_image=str(raw.get("source_image") or ""),
                    destination=destination,
                    actual_destination=str(actual_destination),
                    status=str(raw.get("status") or ""),
                    destination_exists=actual_destination.exists() if check_files else False,
                )
            )
    return rows


def load_dataset_csv_index(path: Path) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for csv_path in sorted(path.glob("*.csv")):
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                hash_md5 = str(row.get("hash_md5") or "").strip().lower()
                if hash_md5:
                    row["csv_origen"] = str(csv_path)
                    index[hash_md5] = row
    return index


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, json.JSONDecodeError):
        return {}


def nested(data: dict[str, Any], *keys: str, default: Any = "") -> Any:
    current: Any = data
    for key in keys:
        if not isinstance(current, dict) or key not in current:
            return default
        current = current[key]
    return current


def number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(result) or math.isinf(result):
        return None
    return result


def avg(values: Iterable[Any]) -> float | None:
    nums = [float(value) for value in values if number(value) is not None]
    return round(statistics.mean(nums), 6) if nums else None


class ReferenceCache:
    def __init__(self, root: Path, deep_image_json: bool = False) -> None:
        self.root = root
        self.deep_image_json = deep_image_json
        self.cache: dict[str, dict[str, Any]] = {}

    def get(self, hash_md5: str) -> dict[str, Any]:
        if hash_md5 in self.cache:
            return self.cache[hash_md5]

        hash_dir = self.root / hash_md5
        metadata_path = hash_dir / "metadata.json"
        static_path = hash_dir / "analysis" / "static_analysis.json"
        image_path = hash_dir / "analysis" / "image_analysis.json"
        metadata = read_json(metadata_path)
        static = read_json(static_path)
        image = read_json(image_path) if self.deep_image_json else {}
        image_summary = image.get("summary") if isinstance(image.get("summary"), dict) else {}
        metadata_image_summary = metadata.get("image_analysis") if isinstance(metadata.get("image_analysis"), dict) else {}
        if not image_summary:
            image_summary = metadata_image_summary

        byte_profile = static.get("byte_profile") if isinstance(static.get("byte_profile"), dict) else {}
        static_hashes = static.get("hashes") if isinstance(static.get("hashes"), dict) else {}
        metadata_static = metadata.get("static_analysis") if isinstance(metadata.get("static_analysis"), dict) else {}
        per_algorithm = image_summary.get("per_algorithm") if isinstance(image_summary.get("per_algorithm"), dict) else {}

        images_by_algorithm: dict[str, dict[str, Any]] = {}
        for item in image.get("images") or []:
            if isinstance(item, dict) and item.get("algorithm"):
                images_by_algorithm[str(item["algorithm"])] = item

        record = {
            "metadata_path": str(metadata_path),
            "static_json_path": str(static_path),
            "image_json_path": str(image_path),
            "json_metadata_exists": metadata_path.exists(),
            "json_static_exists": static_path.exists(),
            "json_image_exists": image_path.exists(),
            "json_image_read_mode": "completo" if self.deep_image_json else "resumen_metadata",
            "generation_status": metadata.get("status", ""),
            "generation_started_at": metadata.get("started_at", ""),
            "generation_finished_at": metadata.get("finished_at", ""),
            "sha1": static_hashes.get("sha1") or metadata_static.get("sha1", ""),
            "sha256": static_hashes.get("sha256") or metadata.get("binary_sha256") or metadata_static.get("sha256", ""),
            "binary_sha256": metadata.get("binary_sha256") or static_hashes.get("sha256") or metadata_static.get("sha256", ""),
            "binary_size_bytes": byte_profile.get("size_bytes") or metadata_static.get("size_bytes", ""),
            "binary_entropy": byte_profile.get("entropy") or metadata_static.get("entropy", ""),
            "printable_ratio": byte_profile.get("printable_ratio", ""),
            "image_total": image_summary.get("total_images", metadata.get("image_count", "")),
            "image_valid": image_summary.get("valid_images", ""),
            "image_failed": image_summary.get("failed_images", ""),
            "image_width_min": image_summary.get("width_min", ""),
            "image_width_max": image_summary.get("width_max", ""),
            "image_height_min": image_summary.get("height_min", ""),
            "image_height_max": image_summary.get("height_max", ""),
            "image_avg_entropy": image_summary.get("avg_entropy", ""),
            "image_avg_brightness": image_summary.get("avg_brightness", ""),
            "image_avg_contrast": image_summary.get("avg_contrast", ""),
            "image_avg_edge_density": image_summary.get("avg_edge_density", ""),
            "per_algorithm": per_algorithm,
            "images_by_algorithm": images_by_algorithm,
        }
        self.cache[hash_md5] = record
        return record


def image_metric(record: dict[str, Any], algorithm_source: str, metric: str) -> Any:
    per_algorithm = record.get("per_algorithm") or {}
    data = per_algorithm.get(algorithm_source) if isinstance(per_algorithm, dict) else {}
    if isinstance(data, dict):
        return data.get(metric, "")
    return ""


def image_item_field(record: dict[str, Any], algorithm_source: str, field: str) -> Any:
    items = record.get("images_by_algorithm") or {}
    data = items.get(algorithm_source) if isinstance(items, dict) else {}
    if isinstance(data, dict):
        return data.get(field, "")
    return ""


def is_available(row: ManifestRow) -> bool:
    return row.status in AVAILABLE_STATUSES


def percent(part: int, total: int) -> float:
    return round((part / total) * 100, 4) if total else 0.0


def split_sort_key(split: str) -> int:
    return SPLIT_ORDER.get(split, 99)


def build_detail_rows(
    rows: list[ManifestRow],
    csv_index: dict[str, dict[str, Any]],
    references: ReferenceCache,
) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    for row in rows:
        source = csv_index.get(row.hash_md5, {})
        ref = references.get(row.hash_md5)
        details.append(
            {
                "split": row.split,
                "family": row.family,
                "hash_md5": row.hash_md5,
                "sha1": ref.get("sha1") or source.get("sha1", ""),
                "sha256": ref.get("sha256") or source.get("sha256", ""),
                "binary_sha256": ref.get("binary_sha256", ""),
                "lote_origen": source.get("lote_origen", ""),
                "tipo_probable": source.get("tipo_probable", ""),
                "detection_percent": number(source.get("detection_percent")),
                "fecha_escaneo_vt": source.get("fecha_escaneo_vt", ""),
                "dia_creacion_archivo": source.get("dia_creacion_archivo", row.creation_day),
                "status_imagen": row.status,
                "archivo_destino_existe": row.destination_exists,
                "binary_size_bytes": number(ref.get("binary_size_bytes")),
                "binary_entropy": number(ref.get("binary_entropy")),
                "printable_ratio": number(ref.get("printable_ratio")),
                "generation_status": ref.get("generation_status", ""),
                "generation_started_at": ref.get("generation_started_at", ""),
                "generation_finished_at": ref.get("generation_finished_at", ""),
                "image_total": ref.get("image_total", ""),
                "image_valid": ref.get("image_valid", ""),
                "image_failed": ref.get("image_failed", ""),
                "alg_image_size_bytes": number(image_item_field(ref, row.algorithm_source, "size_bytes")),
                "alg_image_sha256": image_item_field(ref, row.algorithm_source, "sha256"),
                "alg_image_format": image_item_field(ref, row.algorithm_source, "format"),
                "alg_image_mode": image_item_field(ref, row.algorithm_source, "mode"),
                "alg_image_width": number(image_item_field(ref, row.algorithm_source, "width")),
                "alg_image_height": number(image_item_field(ref, row.algorithm_source, "height")),
                "alg_image_pixels": number(image_item_field(ref, row.algorithm_source, "pixels")),
                "alg_image_aspect_ratio": number(image_item_field(ref, row.algorithm_source, "aspect_ratio")),
                "alg_avg_entropy": number(image_metric(ref, row.algorithm_source, "avg_entropy")),
                "alg_avg_brightness": number(image_metric(ref, row.algorithm_source, "avg_brightness")),
                "alg_avg_contrast": number(image_metric(ref, row.algorithm_source, "avg_contrast")),
                "alg_avg_edge_density": number(image_metric(ref, row.algorithm_source, "avg_edge_density")),
                "source_image": row.source_image,
                "destination_manifest": row.destination,
                "destination_real": row.actual_destination,
                "metadata_json": ref.get("metadata_path", ""),
                "static_json": ref.get("static_json_path", ""),
                "image_json": ref.get("image_json_path", ""),
                "json_metadata_exists": ref.get("json_metadata_exists", False),
                "json_static_exists": ref.get("json_static_exists", False),
                "json_image_exists": ref.get("json_image_exists", False),
            }
        )
    return details


def counter_rows(rows: list[dict[str, Any]], fields: list[str]) -> list[dict[str, Any]]:
    counter: Counter[tuple[Any, ...]] = Counter(tuple(row.get(field, "") for field in fields) for row in rows)
    total = sum(counter.values())
    output = []
    for key, count in sorted(counter.items(), key=lambda item: tuple(str(v) for v in item[0])):
        row = {field: value for field, value in zip(fields, key)}
        row["muestras"] = count
        row["porcentaje"] = percent(count, total)
        output.append(row)
    return output


def split_family_summary(details: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in details:
        grouped[(str(row["split"]), str(row["family"]))].append(row)

    output = []
    for (split, family), items in sorted(grouped.items(), key=lambda item: (split_sort_key(item[0][0]), item[0][1])):
        available = sum(1 for row in items if row["status_imagen"] in AVAILABLE_STATUSES)
        missing = sum(1 for row in items if row["status_imagen"] == "imagen_no_encontrada")
        dates = [row["dia_creacion_archivo"] for row in items if row.get("dia_creacion_archivo")]
        output.append(
            {
                "split": split,
                "family": family,
                "muestras": len(items),
                "disponibles": available,
                "faltantes": missing,
                "cobertura_pct": percent(available, len(items)),
                "fecha_min": min(dates) if dates else "",
                "fecha_max": max(dates) if dates else "",
                "deteccion_prom": avg(row.get("detection_percent") for row in items),
                "binary_size_prom": avg(row.get("binary_size_bytes") for row in items),
                "binary_entropy_prom": avg(row.get("binary_entropy") for row in items),
                "printable_ratio_prom": avg(row.get("printable_ratio") for row in items),
                "alg_entropy_prom": avg(row.get("alg_avg_entropy") for row in items),
                "alg_brightness_prom": avg(row.get("alg_avg_brightness") for row in items),
                "alg_contrast_prom": avg(row.get("alg_avg_contrast") for row in items),
                "alg_edge_density_prom": avg(row.get("alg_avg_edge_density") for row in items),
            }
        )
    return output


def month_summary(details: list[dict[str, Any]]) -> list[dict[str, Any]]:
    months = sorted({str(row.get("dia_creacion_archivo", ""))[:7] for row in details if row.get("dia_creacion_archivo")})
    splits = ["train", "validacion", "test"]
    counter: Counter[tuple[str, str]] = Counter()
    for row in details:
        month = str(row.get("dia_creacion_archivo", ""))[:7]
        if month:
            counter[(month, str(row["split"]))] += 1
    output = []
    for month in months:
        record = {"mes": month}
        total = 0
        for split in splits:
            count = counter[(month, split)]
            record[split] = count
            total += count
        record["total"] = total
        output.append(record)
    return output


def metric_summary(details: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in details:
        grouped[(str(row["split"]), str(row["family"]))].append(row)
    output = []
    for (split, family), items in sorted(grouped.items(), key=lambda item: (split_sort_key(item[0][0]), item[0][1])):
        output.append(
            {
                "split": split,
                "family": family,
                "muestras": len(items),
                "binary_size_min": min([v for v in (number(row.get("binary_size_bytes")) for row in items) if v is not None], default=None),
                "binary_size_prom": avg(row.get("binary_size_bytes") for row in items),
                "binary_size_max": max([v for v in (number(row.get("binary_size_bytes")) for row in items) if v is not None], default=None),
                "binary_entropy_prom": avg(row.get("binary_entropy") for row in items),
                "printable_ratio_prom": avg(row.get("printable_ratio") for row in items),
                "image_width_prom": avg(row.get("alg_image_width") for row in items),
                "image_height_prom": avg(row.get("alg_image_height") for row in items),
                "image_pixels_prom": avg(row.get("alg_image_pixels") for row in items),
                "alg_entropy_prom": avg(row.get("alg_avg_entropy") for row in items),
                "alg_brightness_prom": avg(row.get("alg_avg_brightness") for row in items),
                "alg_contrast_prom": avg(row.get("alg_avg_contrast") for row in items),
                "alg_edge_density_prom": avg(row.get("alg_avg_edge_density") for row in items),
            }
        )
    return output


def append_rows(ws, rows: list[dict[str, Any]], headers: list[str] | None = None) -> None:
    if headers is None:
        headers = []
        for row in rows:
            for key in row:
                if key not in headers:
                    headers.append(key)
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    set_column_widths(ws, headers)


def style_header(ws, row_index: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_index, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 10), 28)
        if header in {"source_image", "destination_manifest", "destination_real", "metadata_json", "static_json", "image_json"}:
            width = 18
        ws.column_dimensions[get_column_letter(index)].width = width


def add_kv(ws, pairs: list[tuple[str, Any]], start_row: int = 1) -> None:
    for row_offset, (key, value) in enumerate(pairs):
        row = start_row + row_offset
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = BOLD_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 38


def add_bar_chart(ws, title: str, data_min_col: int, data_max_col: int, min_row: int, max_row: int, cats_col: int, anchor: str) -> None:
    if max_row <= min_row:
        return
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Muestras"
    chart.x_axis.title = "Categoria"
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=cats_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 13
    ws.add_chart(chart, anchor)


def add_line_chart(ws, title: str, data_min_col: int, data_max_col: int, min_row: int, max_row: int, cats_col: int, anchor: str) -> None:
    if max_row <= min_row:
        return
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Muestras"
    chart.x_axis.title = "Mes"
    data = Reference(ws, min_col=data_min_col, max_col=data_max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=cats_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 18
    ws.add_chart(chart, anchor)


def add_pie_chart(ws, title: str, labels_col: int, data_col: int, min_row: int, max_row: int, anchor: str) -> None:
    if max_row < min_row:
        return
    chart = PieChart()
    chart.title = title
    labels = Reference(ws, min_col=labels_col, min_row=min_row, max_row=max_row)
    data = Reference(ws, min_col=data_col, min_row=min_row, max_row=max_row)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.height = 7
    chart.width = 9
    ws.add_chart(chart, anchor)


def create_summary_sheet(wb: Workbook, dataset_name: str, rows: list[ManifestRow], details: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "Resumen"
    available = sum(1 for row in rows if is_available(row))
    missing = sum(1 for row in rows if row.status == "imagen_no_encontrada")
    exists = sum(1 for row in rows if row.destination_exists)
    families = sorted({row.family for row in rows})
    dates = [row.creation_day for row in rows if row.creation_day]
    ref_missing = sum(1 for row in details if not row["json_metadata_exists"] or not row["json_static_exists"] or not row["json_image_exists"])
    add_kv(
        ws,
        [
            ("dataset_folder_manifest", rows[0].dataset_folder),
            ("dataset_folder_real", rows[0].actual_dataset_folder),
            ("balance", rows[0].balance_sheet),
            ("balance_nombre", rows[0].balance_name),
            ("algoritmo", rows[0].algorithm),
            ("algoritmo_fuente", rows[0].algorithm_source),
            ("muestras_esperadas", len(rows)),
            ("muestras_disponibles", available),
            ("muestras_faltantes", missing),
            ("cobertura_pct", percent(available, len(rows))),
            ("archivos_existentes", exists),
            ("hashes_unicos", len({row.hash_md5 for row in rows})),
            ("familias", len(families)),
            ("familias_lista", ", ".join(families)),
            ("fecha_min", min(dates) if dates else ""),
            ("fecha_max", max(dates) if dates else ""),
            ("referencias_json_incompletas", ref_missing),
        ],
    )

    status_rows = [{"estado": status, "muestras": count} for status, count in sorted(Counter(row.status for row in rows).items())]
    start = 22
    ws.cell(row=start, column=1, value="Estado imagen")
    ws.cell(row=start, column=2, value="Muestras")
    style_header(ws, start, 2)
    for offset, row in enumerate(status_rows, start=1):
        ws.cell(row=start + offset, column=1, value=row["estado"])
        ws.cell(row=start + offset, column=2, value=row["muestras"])
    add_pie_chart(ws, "Estado de imagenes", 1, 2, start + 1, start + len(status_rows), "D21")

    split_rows = counter_rows(details, ["split"])
    split_start = 22
    ws.cell(row=split_start, column=7, value="Split")
    ws.cell(row=split_start, column=8, value="Muestras")
    style_header(ws, split_start, 8)
    for offset, row in enumerate(split_rows, start=1):
        ws.cell(row=split_start + offset, column=7, value=row["split"])
        ws.cell(row=split_start + offset, column=8, value=row["muestras"])
    add_bar_chart(ws, "Distribucion por split", 8, 8, split_start, split_start + len(split_rows), 7, "J21")


def create_workbook(dataset_name: str, rows: list[ManifestRow], csv_index: dict[str, dict[str, Any]], references: ReferenceCache) -> Workbook:
    rows = sorted(rows, key=lambda row: (split_sort_key(row.split), row.family, row.creation_day, row.hash_md5))
    details = build_detail_rows(rows, csv_index, references)
    wb = Workbook()
    create_summary_sheet(wb, dataset_name, rows, details)

    split_family = split_family_summary(details)
    ws = wb.create_sheet("Split_Familia")
    append_rows(ws, split_family)
    add_bar_chart(ws, "Muestras por split/familia", 3, 3, 1, min(ws.max_row, 25), 2, "Q2")
    add_bar_chart(ws, "Cobertura por split/familia", 6, 6, 1, min(ws.max_row, 25), 2, "Q18")

    metrics = metric_summary(details)
    ws = wb.create_sheet("Metricas")
    append_rows(ws, metrics)
    add_bar_chart(ws, "Entropia binaria promedio", 8, 8, 1, min(ws.max_row, 25), 2, "P2")
    add_bar_chart(ws, "Entropia de imagen promedio", 12, 12, 1, min(ws.max_row, 25), 2, "P18")

    monthly = month_summary(details)
    ws = wb.create_sheet("Fechas")
    append_rows(ws, monthly, ["mes", "train", "validacion", "test", "total"])
    add_line_chart(ws, "Muestras por mes y split", 2, 5, 1, ws.max_row, 1, "G2")

    missing = [
        {
            "split": row["split"],
            "family": row["family"],
            "hash_md5": row["hash_md5"],
            "dia_creacion_archivo": row["dia_creacion_archivo"],
            "status_imagen": row["status_imagen"],
            "source_image": row["source_image"],
            "destination_real": row["destination_real"],
            "image_json": row["image_json"],
        }
        for row in details
        if row["status_imagen"] == "imagen_no_encontrada"
    ]
    ws = wb.create_sheet("Faltantes")
    append_rows(ws, missing or [{"mensaje": "Sin faltantes"}])

    ws = wb.create_sheet("Muestras")
    detail_headers = [
        "split",
        "family",
        "hash_md5",
        "sha1",
        "sha256",
        "binary_sha256",
        "lote_origen",
        "tipo_probable",
        "detection_percent",
        "fecha_escaneo_vt",
        "dia_creacion_archivo",
        "status_imagen",
        "archivo_destino_existe",
        "binary_size_bytes",
        "binary_entropy",
        "printable_ratio",
        "generation_status",
        "generation_started_at",
        "generation_finished_at",
        "image_total",
        "image_valid",
        "image_failed",
        "alg_image_size_bytes",
        "alg_image_sha256",
        "alg_image_format",
        "alg_image_mode",
        "alg_image_width",
        "alg_image_height",
        "alg_image_pixels",
        "alg_image_aspect_ratio",
        "alg_avg_entropy",
        "alg_avg_brightness",
        "alg_avg_contrast",
        "alg_avg_edge_density",
        "source_image",
        "destination_real",
        "metadata_json",
        "static_json",
        "image_json",
        "json_metadata_exists",
        "json_static_exists",
        "json_image_exists",
    ]
    append_rows(ws, details, detail_headers)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")
    return wb


def group_by_dataset(rows: list[ManifestRow]) -> dict[str, list[ManifestRow]]:
    grouped: defaultdict[str, list[ManifestRow]] = defaultdict(list)
    for row in rows:
        grouped[row.dataset_folder].append(row)
    return dict(sorted(grouped.items()))


def save_workbooks(args: argparse.Namespace) -> list[Path]:
    manifest_rows = load_manifest(args.manifest, check_files=not args.no_file_check)
    csv_index = load_dataset_csv_index(args.dataset_csv_dir)
    references = ReferenceCache(args.json_root, deep_image_json=args.deep_image_json)
    grouped = group_by_dataset(manifest_rows)
    selected = set(args.datasets or [])
    outputs: list[Path] = []

    args.output_dir.mkdir(parents=True, exist_ok=True)
    for dataset_folder, rows in grouped.items():
        actual_folder = rows[0].actual_dataset_folder
        if selected and dataset_folder not in selected and actual_folder not in selected:
            continue
        filename = f"{safe_name(actual_folder)}.xlsx"
        output_path = args.output_dir / filename
        print(f"Generando {output_path} ({len(rows):,} muestras)...")
        wb = create_workbook(dataset_folder, rows, csv_index, references)
        wb.save(output_path)
        outputs.append(output_path)
    return outputs


def main() -> None:
    args = parse_args()
    outputs = save_workbooks(args)
    print("Excel por dataset generados")
    print(f"Total workbooks: {len(outputs):,}")
    for path in outputs:
        print(path.resolve())


if __name__ == "__main__":
    main()

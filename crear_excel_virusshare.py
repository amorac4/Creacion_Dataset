#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Consolida los CSV filtrados de VirusShare en un solo archivo XLSX."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill


DEFAULT_INPUT_DIR = r"dataset_filtrado\VirusShare_00499"
DEFAULT_OUTPUT = r"outputs\VirusShare_00499_dataset_filtrado.xlsx"
EXCEL_MAX_ROWS = 1_048_576
DETECTIONS_ROWS_PER_SHEET = 1_000_000

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Crea un Excel consolidado desde los CSV filtrados de VirusShare."
    )
    parser.add_argument("--input-dir", default=DEFAULT_INPUT_DIR, help="Carpeta con los CSV filtrados.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Ruta del archivo XLSX final.")
    return parser.parse_args()


def styled_row(ws, values: Iterable[str], is_header: bool = False) -> list[WriteOnlyCell | str]:
    if not is_header:
        return list(values)
    cells: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(ws, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cells.append(cell)
    return cells


def set_sheet_basics(ws, widths: dict[str, int]) -> None:
    ws.freeze_panes = "A2"
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def append_csv_to_sheet(wb: Workbook, csv_path: Path, sheet_name: str, widths: dict[str, int]) -> int:
    ws = wb.create_sheet(title=sheet_name)
    set_sheet_basics(ws, widths)
    row_count = 0
    with csv_path.open("r", encoding="utf-8", newline="") as file:
        reader = csv.reader(file)
        for row in reader:
            ws.append(styled_row(ws, row, is_header=(row_count == 0)))
            row_count += 1
    return max(row_count - 1, 0)


def append_detections_split(wb: Workbook, csv_path: Path) -> tuple[int, int]:
    total_rows = 0
    sheet_count = 0
    current_ws = None
    current_data_rows = 0
    header: list[str] | None = None
    widths = {"A": 34, "B": 22, "C": 48, "D": 20, "E": 16, "F": 24}

    with csv_path.open("r", encoding="utf-8", newline="") as file:
        reader = csv.reader(file)
        for row_number, row in enumerate(reader):
            if row_number == 0:
                header = row
                continue

            if current_ws is None or current_data_rows >= DETECTIONS_ROWS_PER_SHEET:
                sheet_count += 1
                current_ws = wb.create_sheet(title=f"Detecciones_{sheet_count}")
                set_sheet_basics(current_ws, widths)
                current_ws.append(styled_row(current_ws, header or [], is_header=True))
                current_data_rows = 0

            current_ws.append(row)
            current_data_rows += 1
            total_rows += 1

    return total_rows, sheet_count


def add_summary_sheet(wb: Workbook, metrics: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title="Resumen", index=0)
    ws.append(styled_row(ws, ["Campo", "Valor"], is_header=True))
    for key, value in metrics:
        ws.append([key, value])
    ws.append([])
    ws.append(["Notas"])
    ws.append(["Las detecciones por motor se dividen en varias hojas por el limite de filas de Excel."])
    ws.append(["La familia y el tipo probable son inferencias heuristicas desde etiquetas AV."])
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 70


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    required = {
        "muestras": input_dir / "muestras_filtradas.csv",
        "detecciones": input_dir / "detecciones_por_motor.csv",
        "familias": input_dir / "resumen_familias.csv",
        "tipos": input_dir / "resumen_tipos.csv",
        "dias": input_dir / "resumen_dias.csv",
        "corruptos": input_dir / "reportes_corruptos.csv",
    }
    missing = [str(path) for path in required.values() if not path.is_file()]
    if missing:
        raise SystemExit("Faltan CSV requeridos:\n" + "\n".join(missing))

    wb = Workbook(write_only=True)
    muestras = append_csv_to_sheet(
        wb,
        required["muestras"],
        "Muestras",
        {"A": 34, "B": 42, "C": 64, "H": 24, "J": 24, "P": 20, "S": 16, "V": 70},
    )
    detecciones, detection_sheets = append_detections_split(wb, required["detecciones"])
    familias = append_csv_to_sheet(wb, required["familias"], "Resumen_Familias", {"A": 26, "B": 14})
    tipos = append_csv_to_sheet(wb, required["tipos"], "Resumen_Tipos", {"A": 22, "B": 14})
    dias = append_csv_to_sheet(wb, required["dias"], "Resumen_Dias", {"A": 18, "B": 14})
    corruptos = append_csv_to_sheet(wb, required["corruptos"], "Reportes_Corruptos", {"A": 80, "B": 60})

    add_summary_sheet(
        wb,
        [
            ("Muestras filtradas", f"{muestras:,}"),
            ("Detecciones por motor", f"{detecciones:,}"),
            ("Hojas de detecciones", str(detection_sheets)),
            ("Familias resumidas", f"{familias:,}"),
            ("Tipos resumidos", f"{tipos:,}"),
            ("Dias resumidos", f"{dias:,}"),
            ("Reportes corruptos", f"{corruptos:,}"),
            ("Limite filas Excel por hoja", f"{EXCEL_MAX_ROWS:,}"),
        ],
    )

    wb.save(output_path)
    print(output_path.resolve())


if __name__ == "__main__":
    main()
